# -*- coding: utf-8 -*-
"""Волна F — отчётность по парку:
261/262 справочник models.json (канон, тип, эталонная прошивка) + отстающие,
263 возраст парка по серийникам NS/PS-YYMMDD, 278 локации из имён (на лету,
через bot_zones), 279 /coverage — матрица корпус×этаж, 280 /floors —
эмодзи-теплокарта, 281 /models — срез по моделям, 282 /cablelog — кабельный
журнал xlsx, 283/284 /report_xlsx — xlsx с диаграммами (PDF без внешних
зависимостей нет — частично), 286 /smis — CSV(UTF-8 BOM)+JSON по стабильной
схеме, 289 /pnr — отчёт ПНР по листу «Изменённые». Всё read-only."""
import io
import os
import re
import csv
import json
import time
import datetime
import tempfile
import collections

import bot_state as st
import bot_inventory as inv
import bot_dq as dq
import bot_store as store
from bot_tg import send, send_chunks, send_document, chat_action
from bot_util import log_exc, esc

SER_RE = re.compile(r"^[A-Z]{2}(\d{2})(\d{2})(\d{2})\d+$")


# ---------- 263: возраст по серийникам ----------
def parse_serial_date(s):
    """'NS220629012790001' -> date(2022,6,29); None — не по формату."""
    m = SER_RE.match(str(s or "").strip().upper())
    if not m:
        return None
    y, mo, d = 2000 + int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        return datetime.date(y, mo, d)
    except ValueError:
        return None


def fleet_age() -> dict:
    """{'switches': [(имя, date)], 'cams': [(ip, date)], 'by_year': Counter}."""
    data = dq.read_all()
    sw, cams = [], []
    d = data.get("Лист1") or {"hdr": [], "rows": []}
    for row in d["rows"]:
        dt = parse_serial_date(dq._cell(row, d["hdr"], "Серийник"))
        if dt:
            sw.append((str(dq._cell(row, d["hdr"], "Название коммутатора")
                           or "?"), dt))
    d = data.get("Изменённые") or {"hdr": [], "rows": []}
    for row in d["rows"]:
        dt = parse_serial_date(dq._cell(row, d["hdr"], "Серийник"))
        if dt:
            cams.append((str(dq._cell(row, d["hdr"], "Новый IP") or "?"), dt))
    for ip, e in inv.fw_cache().items():
        dt = parse_serial_date(e.get("serial"))
        if dt:
            cams.append((ip, dt))
    by_year = collections.Counter(d.year for _n, d in sw + cams)
    return {"switches": sw, "cams": cams, "by_year": by_year}


# ---------- 261/262: models.json ----------
def _model_type(name: str) -> str:
    low = str(name).lower()
    for pat, t in (("bullet", "bullet"), ("dome", "dome"), ("ptz", "PTZ"),
                   ("box", "box")):
        if pat in low:
            return t
    return "?"


def refresh_models() -> dict:
    """261: автосправочник моделей + 262 эталонная прошивка (самая частая
    из ONVIF-кэша; ручные правки models.json сохраняются)."""
    path = st.cget("models_path")
    ref = store.jload(path, {})
    cnt = collections.Counter()
    for c in inv.cams():
        if c.get("model"):
            cnt[str(c["model"]).strip()] += 1
    fw_by_model = collections.defaultdict(collections.Counter)
    for e in inv.fw_cache().values():
        if e.get("model") and e.get("fw"):
            fw_by_model[str(e["model"]).strip()][e["fw"]] += 1
    mm = dq.model_map()
    canon_cnt = collections.Counter()
    for name, n in cnt.items():
        canon_cnt[mm.get(dq._mkey(name), name)] += n
    for canon, n in canon_cnt.items():
        e = ref.setdefault(canon, {"type": _model_type(canon)})
        e["count"] = n
        if not e.get("ref_fw_manual"):
            for mname, fws in fw_by_model.items():
                if dq._mkey(mname) and dq._mkey(mname) in dq._mkey(canon):
                    e["ref_fw"] = fws.most_common(1)[0][0]
                    break
    try:
        store.jsave(path, ref)
    except Exception:
        log_exc("reports: models.json не сохранился")
    return ref


def fw_lag() -> list:
    """262: [(ip, model, fw, ref_fw)] — камеры не на эталонной прошивке."""
    ref = store.jload(st.cget("models_path"), {})
    out = []
    for ip, e in inv.fw_cache().items():
        for canon, meta in ref.items():
            if meta.get("ref_fw") and e.get("model") \
                    and dq._mkey(e["model"]) in dq._mkey(canon) \
                    and e.get("fw") and e["fw"] != meta["ref_fw"]:
                out.append((ip, e["model"], e["fw"], meta["ref_fw"]))
    return out


# ---------- 281 (+263 +262): /models ----------
def cmd_models(chat, arg="", reply_to=None):
    chat_action(chat)
    cams = [c for c in inv.cams() if c.get("ip")]
    models = collections.Counter(str(c.get("model") or "—").strip() for c in cams)
    hm = _health()
    ref = refresh_models()
    lines = [f"🎥 <b>Срез по моделям</b> — {len(cams)} камер, "
             f"{len(models)} написаний ({len(ref)} канонов):"]
    for m, n in models.most_common(15):
        on = sum(1 for c in cams if str(c.get("model") or "—").strip() == m
                 and (hm.get(c["ip"]) or {}).get("ok"))
        t = (ref.get(m) or {}).get("type", "?")
        lines.append(f"• {esc(m)} [{esc(t)}]: <b>{n}</b> "
                     f"({100 * n // len(cams)}%) · 🟢 {on}")
    varn = dq.model_variants()
    if varn:
        lines.append(f"\n✍️ Разнобой написаний ({len(varn)}), канон → варианты:")
        for canon, vs in list(varn.items())[:6]:
            lines.append(f"• {esc(canon)} ⇐ {esc(' | '.join(v for v in vs if v != canon))}")
    lag = fw_lag()
    if lag:
        lines.append(f"\n🧩 Не на эталонной прошивке: <b>{len(lag)}</b> "
                     f"(эталон = самая частая в ONVIF-кэше):")
        for ip, m, fw, rfw in lag[:8]:
            lines.append(f"• <code>{ip}</code> {esc(fw)} ≠ {esc(rfw)}")
    age = fleet_age()
    if age["by_year"]:
        lines.append("\n📅 Возраст по серийникам (NS/PS-YYMMDD…): "
                     + " · ".join(f"{y}: {n}"
                                  for y, n in sorted(age["by_year"].items())))
        lines.append(f"(камер с датируемым серийником: {len(age['cams'])}, "
                     f"коммутаторов: {len(age['switches'])}; серийники Apix — "
                     f"хвост MAC, даты не содержат)")
    send_chunks(chat, lines)


# ---------- 278/279/280: покрытие ----------
def _health() -> dict:
    try:
        import bot_health as bh
        return bh.snapshot()["ips"]
    except Exception:
        return {}


def coverage_matrix() -> dict:
    """279: {(корпус, этаж): [total, online]} по 278-разбору имён/локаций."""
    import bot_zones
    hm = _health()
    mat = collections.defaultdict(lambda: [0, 0])
    for c in inv.cams():
        if not c.get("ip"):
            continue
        m = bot_zones.cam_meta(c)
        key = (m.get("bld") or "—", m.get("floor")
               if m.get("floor") is not None else "?")
        mat[key][0] += 1
        if (hm.get(c["ip"]) or {}).get("ok"):
            mat[key][1] += 1
    return dict(mat)


def cmd_coverage(chat, arg="", reply_to=None):
    mat = coverage_matrix()
    if not mat:
        send(chat, "Матрица покрытия пуста — нет камер с IP.", reply_to=reply_to)
        return
    blds = sorted({k[0] for k in mat})
    floors = sorted({k[1] for k in mat},
                    key=lambda f: (isinstance(f, str), f))
    rows = ["эт\\крп " + " ".join(f"{str(b)[:4]:>5}" for b in blds)]
    for fl in floors:
        cells = []
        for b in blds:
            t, on = mat.get((b, fl), (0, 0))
            cells.append(f"{on}/{t}" if t else "  ·  ")
        rows.append(f"{str(fl):>6} " + " ".join(f"{c:>5}" for c in cells))
    holes = [f"{b}/{fl}" for (b, fl), (t, on) in sorted(mat.items(),
                                                        key=lambda x: str(x[0]))
             if t and on == 0]
    txt = (f"🗺 <b>Покрытие корпус × этаж</b> (онлайн/всего):\n"
           f"<pre>{esc(chr(10).join(rows))}</pre>")
    if holes:
        txt += ("\n🕳 Дыры (всё офлайн): "
                + ", ".join(esc(h) for h in holes[:15]))
    txt += "\n«—» = имя не по схеме (278: /names), «?» = этаж не распознан."
    send(chat, txt, reply_to=reply_to)


def cmd_floors(chat, arg="", reply_to=None):
    """280: эмодзи-теплокарта этажей по доле онлайн."""
    import bot_zones
    hm = _health()
    per = collections.defaultdict(lambda: [0, 0])
    for c in inv.cams():
        if not c.get("ip"):
            continue
        fl = bot_zones.floor_of(c)
        key = fl if fl is not None else "?"
        per[key][0] += 1
        if (hm.get(c["ip"]) or {}).get("ok"):
            per[key][1] += 1
    if not per:
        send(chat, "Нет данных по этажам.", reply_to=reply_to)
        return
    lines = ["🌡 <b>Теплокарта этажей</b> (доля онлайн):"]
    for fl in sorted(per, key=lambda f: (isinstance(f, str), f)):
        t, on = per[fl]
        share = on / t if t else 0
        mark = "🟩" if share >= 0.95 else ("🟨" if share >= 0.8 else "🟥")
        bar = "▰" * round(share * 10) + "▱" * (10 - round(share * 10))
        lines.append(f"{mark} эт.{str(fl):>3}: {bar} {on}/{t}")
    if not _health():
        lines.append("❔ health-check ещё не бегал — всё как офлайн.")
    send(chat, "\n".join(lines), reply_to=reply_to)


# ---------- 282: /cablelog ----------
def cmd_cablelog(chat, arg="", reply_to=None, out_path=None):
    chat_action(chat, "upload_document")
    import openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Кабельный журнал"
    hdr = ["№", "Коммутатор", "IP коммутатора", "Порт", "VLAN",
           "Длина (м)", "Камера", "IP камеры", "MAC", "Локация", "Объект"]
    ws.append(hdr)
    for c in ws[1]:
        c.font = Font(bold=True)
    cams = sorted([c for c in inv.cams() if c.get("sw_ip") or c.get("switch")],
                  key=lambda r: (str(r.get("sw_ip") or ""),
                                 int(re.sub(r"\D", "", str(r.get("port") or "0"))
                                     or 0)))
    for i, c in enumerate(cams, 1):
        ws.append([i, c.get("switch"), c.get("sw_ip"), c.get("port"),
                   c.get("vlan"), c.get("cable"), c.get("name"), c.get("ip"),
                   c.get("mac"), c.get("location"), c.get("obj")])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{len(cams) + 1}"
    for col, w in zip("ABCDEFGHIJK", (5, 16, 14, 7, 6, 9, 14, 14, 19, 24, 14)):
        ws.column_dimensions[col].width = w
    if out_path:
        wb.save(out_path)
        return out_path
    fd, tmp = tempfile.mkstemp(suffix=".xlsx", prefix="cable_")
    os.close(fd)
    try:
        wb.save(tmp)
        with open(tmp, "rb") as f:
            data = f.read()
    finally:
        try:
            os.unlink(tmp)
        except OSError:
            pass
    send_document(chat, data,
                  f"Кабельный_журнал_{time.strftime('%Y%m%d')}.xlsx",
                  caption=f"🧵 Кабельный журнал: {len(cams)} линий "
                          f"(для подрядчиков)")


# ---------- 283/284: /report_xlsx ----------
def cmd_report_xlsx(chat, arg="", reply_to=None, out_path=None):
    """284 (283 частично: xlsx вместо PDF — без внешних зависимостей)."""
    chat_action(chat, "upload_document")
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.chart import PieChart, BarChart, Reference
    hm = _health()
    cams = [c for c in inv.cams() if c.get("ip")]
    online = sum(1 for c in cams if (hm.get(c["ip"]) or {}).get("ok"))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KPI"
    ws.append(["Executive-отчёт по парку камер МФК «Зарядье»",
               time.strftime("%Y-%m-%d %H:%M")])
    ws["A1"].font = Font(bold=True, size=14)
    for k, v in (("Камер в инвентаре", len(cams)),
                 ("Онлайн (health)", online),
                 ("Офлайн", len(cams) - online if hm else "нет данных"),
                 ("DQ-score", "")):
        ws.append([k, v])
    try:
        scores = dq.dq_scores()
        ws["B5"] = round(sum(s for s, _ in scores) / max(len(scores), 1), 1)
    except Exception:
        log_exc("report_xlsx: dq")
    # модели + pie
    ws.append([])
    ws.append(["Модель", "Камер"])
    models = collections.Counter(str(c.get("model") or "—") for c in cams)
    r0 = ws.max_row
    for m, n in models.most_common(10):
        ws.append([m, n])
    pie = PieChart()
    pie.title = "Камеры по моделям"
    pie.add_data(Reference(ws, min_col=2, min_row=r0 + 1,
                           max_row=ws.max_row), titles_from_data=False)
    pie.set_categories(Reference(ws, min_col=1, min_row=r0 + 1,
                                 max_row=ws.max_row))
    ws.add_chart(pie, "D3")
    # подсети + bar
    ws.append([])
    ws.append(["Подсеть", "Всего", "Онлайн"])
    r1 = ws.max_row
    by_sub = collections.defaultdict(lambda: [0, 0])
    for c in cams:
        sub = c["ip"].rsplit(".", 1)[0]
        by_sub[sub][0] += 1
        if (hm.get(c["ip"]) or {}).get("ok"):
            by_sub[sub][1] += 1
    for sub, (t, on) in sorted(by_sub.items()):
        ws.append([sub + ".x", t, on])
    bar = BarChart()
    bar.title = "Всего/онлайн по подсетям"
    bar.add_data(Reference(ws, min_col=2, max_col=3, min_row=r1,
                           max_row=ws.max_row), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=1, min_row=r1 + 1,
                                 max_row=ws.max_row))
    ws.add_chart(bar, "D20")
    # лист «Проблемы» + условная заливка офлайн
    ws2 = wb.create_sheet("Проблемы")
    ws2.append(["IP", "Название", "Локация", "Свитч", "Порт", "Статус"])
    red = PatternFill("solid", fgColor="F5C6C6")
    for c in cams:
        off = hm and (hm.get(c["ip"]) or {}).get("ok") is False
        stale = "офлайн" in str(c.get("status") or "").lower()
        if off or stale:
            ws2.append([c["ip"], c.get("name"), c.get("location"),
                        c.get("switch"), c.get("port"),
                        "офлайн (health)" if off else str(c.get("status"))])
            for cell in ws2[ws2.max_row]:
                cell.fill = red
    if out_path:
        wb.save(out_path)
        return out_path
    fd, tmp = tempfile.mkstemp(suffix=".xlsx", prefix="report_")
    os.close(fd)
    try:
        wb.save(tmp)
        with open(tmp, "rb") as f:
            data = f.read()
    finally:
        try:
            os.unlink(tmp)
        except OSError:
            pass
    send_document(chat, data, f"Отчёт_парк_{time.strftime('%Y%m%d')}.xlsx",
                  caption=f"📈 Отчёт: {online}/{len(cams)} онлайн, диаграммы, "
                          f"лист «Проблемы» (PDF → xlsx, 283 частично)")


# ---------- 286: /smis ----------
SMIS_FIELDS = ("uid", "ip", "name", "model", "location", "object",
               "switch", "switch_ip", "port", "vlan", "status", "dq_score")


def smis_rows() -> list:
    scores = {id(rec): s for s, rec in dq.dq_scores()}
    out = []
    for c in inv.cams():
        out.append({"uid": dq.mac_canon(c.get("mac")) or c.get("nmac") or "",
                    "ip": c.get("ip") or "", "name": str(c.get("name") or ""),
                    "model": str(c.get("model") or ""),
                    "location": str(c.get("location") or ""),
                    "object": str(c.get("obj") or ""),
                    "switch": str(c.get("switch") or ""),
                    "switch_ip": str(c.get("sw_ip") or ""),
                    "port": str(c.get("port") or ""),
                    "vlan": str(c.get("vlan") or ""),
                    "status": str(c.get("status") or ""),
                    "dq_score": scores.get(id(c), "")})
    out.sort(key=lambda r: r["ip"])
    return out


def cmd_smis(chat, arg="", reply_to=None):
    chat_action(chat, "upload_document")
    ver = str(st.cget("smis_schema_version"))
    rows = smis_rows()
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";", lineterminator="\r\n")
    w.writerow([f"# schema=smis-cameras v{ver}",
                time.strftime("%Y-%m-%d %H:%M")])
    w.writerow(SMIS_FIELDS)
    for r in rows:
        w.writerow([r[f] for f in SMIS_FIELDS])
    send_document(chat, ("﻿" + buf.getvalue()).encode("utf-8"),
                  f"smis_cameras_{time.strftime('%Y%m%d')}.csv",
                  caption=f"🏛 Экспорт для СМИС · схема v{ver} · {len(rows)} строк")
    js = json.dumps({"schema": f"smis-cameras v{ver}",
                     "generated": time.strftime("%Y-%m-%d %H:%M"),
                     "cameras": rows}, ensure_ascii=False, indent=1)
    send_document(chat, js.encode("utf-8"),
                  f"smis_cameras_{time.strftime('%Y%m%d')}.json",
                  caption="🏛 То же в JSON")


# ---------- 289: /pnr ----------
def cmd_pnr(chat, arg="", reply_to=None):
    try:
        days = max(1, min(int(arg), 365))
    except (TypeError, ValueError):
        days = 30
    d = dq.read_all().get("Изменённые") or {"hdr": [], "rows": []}
    cut = datetime.datetime.now() - datetime.timedelta(days=days)
    n_new = n_ip = 0
    by_sw = collections.Counter()
    by_day = collections.Counter()
    for row in d["rows"]:
        dts = str(dq._cell(row, d["hdr"], "Дата и время") or "")
        try:
            when = datetime.datetime.strptime(dts[:16], "%Y-%m-%d %H:%M")
        except ValueError:
            continue
        if when < cut:
            continue
        what = str(dq._cell(row, d["hdr"], "Что сделано") or "").lower()
        if "заводск" in what:
            n_new += 1
        else:
            n_ip += 1
        by_sw[str(dq._cell(row, d["hdr"], "Коммутатор")
                  or dq._cell(row, d["hdr"], "IP коммутатора") or "?")] += 1
        by_day[when.date().isoformat()] += 1
    total = n_new + n_ip
    if not total:
        send(chat, f"🔧 ПНР за {days} дн.: записей в «Изменённых» нет.",
             reply_to=reply_to)
        return
    lines = [f"🔧 <b>Отчёт ПНР за {days} дн.</b> (лист «Изменённые»)",
             f"Всего операций: <b>{total}</b> · заводских введено: "
             f"<b>{n_new}</b> · прочих смен IP: {n_ip}",
             f"Рабочих дней: {len(by_day)} · средний темп: "
             f"{total / max(len(by_day), 1):.1f}/день"]
    if by_sw:
        lines.append("По коммутаторам: " + " · ".join(
            f"{esc(k)}: {v}" for k, v in by_sw.most_common(8)))
    if by_day:
        lines.append("По дням: " + " · ".join(
            f"{k[5:]}: {v}" for k, v in sorted(by_day.items())[-10:]))
    lines.append(f"Всего в листе за всё время: {len(d['rows'])} записей")
    send_chunks(chat, lines)


HANDLERS = {
    "/models": cmd_models, "/floors": cmd_floors, "/coverage": cmd_coverage,
    "/cablelog": cmd_cablelog, "/report_xlsx": cmd_report_xlsx,
    "/smis": cmd_smis, "/pnr": cmd_pnr,
}
ALIASES = {"/модели": "/models", "/этажи": "/floors", "/покрытие": "/coverage",
           "/кабжур": "/cablelog", "/пнр": "/pnr"}
CALLBACKS = {}
