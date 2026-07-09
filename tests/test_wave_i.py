# -*- coding: utf-8 -*-
"""Волна I (401-450): алгоритм дифф-синка на фикстурах (429/413/435),
md5-дедуп заливок (423), чанкование resumable (445), расписание ночных
тиков (446-450), ротация снапшот-копий (427), строители оформления
(404/407/409), сверка сеть-vs-инвентарь (446). Google и сеть НЕ трогаются
(всё на моках и временных файлах); прод-xlsx не пишется."""
import os
import sys
import datetime
import tempfile
import unittest

sys.path.insert(0, r"C:\Users\1\camera")

import bot_state as st  # noqa: E402
import google_api as g  # noqa: E402
import bot_gsheets2 as gs  # noqa: E402
import bot_gfmt as gf  # noqa: E402
import bot_nightly as ni  # noqa: E402


class TestDiffAlgo(unittest.TestCase):
    def test_col_letter(self):
        self.assertEqual(gs.col_letter(0), "A")
        self.assertEqual(gs.col_letter(25), "Z")
        self.assertEqual(gs.col_letter(26), "AA")
        self.assertEqual(gs.col_letter(27), "AB")

    def test_norm_cell(self):
        self.assertEqual(gs.norm_cell(None), "")
        self.assertEqual(gs.norm_cell(5.0), "5")   # UNFORMATTED float vs int
        self.assertEqual(gs.norm_cell(5), "5")
        self.assertEqual(gs.norm_cell(5.5), "5.5")
        self.assertEqual(gs.norm_cell(True), "TRUE")
        self.assertEqual(gs.norm_cell("10.20.50.51"), "10.20.50.51")

    def test_grid_identical(self):
        rows = [["IP", "Имя"], ["10.20.50.1", "AS-1"]]
        d = gs.grid_diff(rows, rows)
        self.assertEqual(d["cells"], 0)
        self.assertEqual(d["blocks"], [])
        self.assertEqual(d["extra_rows"], 0)

    def test_grid_numeric_equivalence(self):
        # 429: 659 в xlsx == 659.0 из UNFORMATTED_VALUE — НЕ изменение
        d = gs.grid_diff([[659.0, "x"]], [[659, "x"]])
        self.assertEqual(d["cells"], 0)

    def test_grid_changed_and_blocks(self):
        remote = [["h1", "h2"], ["a", "b"], ["c", "d"], ["e", "f"], ["g", "h"]]
        local = [["h1", "h2"], ["a", "B"], ["c", "D"], ["e", "f"], ["g", "H"]]
        d = gs.grid_diff(remote, local)
        self.assertEqual(d["cells"], 3)
        self.assertEqual(d["rows"], 3)
        # строки 1-2 смежные -> один блок, строка 4 -> отдельный
        self.assertEqual(len(d["blocks"]), 2)
        self.assertEqual(d["blocks"][0][0], 1)
        self.assertEqual(len(d["blocks"][0][1]), 2)
        self.assertEqual(d["blocks"][1][0], 4)

    def test_grid_new_rows_and_extra(self):
        remote = [["h"], ["a"], ["b"], ["c"]]
        local = [["h"], ["a"]]
        d = gs.grid_diff(remote, local)
        self.assertEqual(d["extra_rows"], 2)      # лишние строки к очистке
        local2 = [["h"], ["a"], ["b"], ["c"], ["NEW"]]
        d2 = gs.grid_diff(remote, local2)
        self.assertEqual(d2["rows"], 1)           # только добавленная строка
        self.assertEqual(d2["blocks"][0][0], 4)

    def test_width_ignores_manual_columns(self):
        # колонки правее ширины xlsx (411 HYPERLINK, ручные) диффом не трогаются
        remote = [["h1", "h2", "ручная"], ["a", "b", "заметка"]]
        local = [["h1", "h2"], ["a", "b"]]
        d = gs.grid_diff(remote, local)
        self.assertEqual(d["cells"], 0)
        self.assertEqual(d["width"], 2)

    def test_row_details(self):
        remote = [["IP-адрес", "Имя"], ["10.1.1.1", "старое"]]
        local = [["IP-адрес", "Имя"], ["10.1.1.1", "новое"]]
        det = gs.row_details(remote, local, key_col=0)
        self.assertEqual(det, [("10.1.1.1", "Имя", "старое", "новое")])

    def test_pull_comments(self):
        rows = [["IP-адрес", "Комментарий"],
                ["10.1.1.1", "фокус сбит"], ["10.1.1.2", ""], ["", "мусор"]]
        self.assertEqual(gs.pull_comments(rows),
                         [("10.1.1.1", "фокус сбит")])
        self.assertEqual(gs.pull_comments([["IP-адрес"]]), [])
        self.assertEqual(gs.pull_comments([]), [])

    def test_apply_comments_on_copy(self):
        """435: запись комментариев в КОПИЮ xlsx (с бэкапом рядом)."""
        import openpyxl
        import bot_inventory as inv
        tmpd = tempfile.mkdtemp(prefix="wave_i_")
        path = os.path.join(tmpd, "test.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = inv.SHEET_MAIN
        ws.append(["IP-адрес", "Имя"])
        ws.append(["10.1.1.1", "AS-1"])
        wb.save(path)
        n = gs.apply_comments([("10.1.1.1", "проверена")], path)
        self.assertEqual(n, 1)
        wb2 = openpyxl.load_workbook(path)
        ws2 = wb2[inv.SHEET_MAIN]
        self.assertEqual(ws2.cell(row=1, column=3).value, "Комментарий")
        self.assertEqual(ws2.cell(row=2, column=3).value, "проверена")
        wb2.close()
        self.assertTrue(any("_comm.xlsx" in f for f in os.listdir(tmpd)))


class TestResumableChunks(unittest.TestCase):
    def test_chunks(self):
        self.assertEqual(g.chunks(0), [(0, 0)])
        self.assertEqual(g.chunks(10, size=4), [(0, 4), (4, 8), (8, 10)])
        self.assertEqual(g.chunks(8, size=4), [(0, 4), (4, 8)])
        blocks = g.chunks(10 * 1024 * 1024)   # дефолтный чанк 4 МиБ
        self.assertEqual(blocks[0], (0, 4 * 1024 * 1024))
        self.assertEqual(blocks[-1][1], 10 * 1024 * 1024)


class TestMd5Dedup(unittest.TestCase):
    def test_dedup_skips_upload(self):
        """423: найден файл с тем же md5 — заливки нет, индекс обновлён."""
        import bot_gdrive2 as gd
        import bot_sheets
        calls = {}
        old_fl, old_up = gd.files_list, g.upload_resumable
        old_idx = bot_sheets._index_update
        gd.files_list = lambda q, **kw: [{"id": "DUP1", "name": "old.jpg"}]
        g.upload_resumable = lambda *a, **kw: (_ for _ in ()).throw(
            AssertionError("не должно заливаться при дедупе"))
        bot_sheets._index_update = lambda ip, e: calls.setdefault("idx", (ip, e))
        try:
            fid = gd.upload_snapshot2("10.9.9.9", b"JPEGDATA")
        finally:
            gd.files_list, g.upload_resumable = old_fl, old_up
            bot_sheets._index_update = old_idx
        self.assertEqual(fid, "DUP1")
        self.assertEqual(calls["idx"][0], "10.9.9.9")
        self.assertEqual(calls["idx"][1]["id"], "DUP1")

    def test_upload_when_new(self):
        """Нет дубликата — заливается resumable с appProperties (424/445)."""
        import bot_gdrive2 as gd
        import bot_sheets
        seen = {}
        old_fl, old_up = gd.files_list, g.upload_resumable
        old_idx, old_df = bot_sheets._index_update, gd.date_folder
        gd.files_list = lambda q, **kw: []
        gd.date_folder = lambda: "FOLDER_DAY"

        def fake_up(name, data, parents=None, mime="", app_properties=None,
                    **kw):
            seen.update({"name": name, "parents": parents,
                         "props": app_properties})
            return {"id": "NEW1", "thumbnailLink": "https://thumb"}
        g.upload_resumable = fake_up
        bot_sheets._index_update = lambda ip, e: seen.setdefault("idx", e)
        try:
            fid = gd.upload_snapshot2("10.9.9.8", b"NEWDATA")
        finally:
            gd.files_list, g.upload_resumable = old_fl, old_up
            bot_sheets._index_update, gd.date_folder = old_idx, old_df
        self.assertEqual(fid, "NEW1")
        self.assertEqual(seen["parents"], ["FOLDER_DAY"])   # 420
        self.assertEqual(seen["props"]["ip"], "10.9.9.8")   # 424
        self.assertIn("md5", seen["props"])                 # 423
        self.assertEqual(seen["idx"]["thumb"], "https://thumb")  # 422


class TestSnapshotRotation(unittest.TestCase):
    def test_rotation_plan(self):
        """427: последние N + по одной на месяц."""
        import bot_gdrive2 as gd
        files = ([{"id": f"n{i}", "name": f"s{i}",
                   "createdTime": f"2026-07-0{9 - i}T00:00:00Z"}
                  for i in range(3)]
                 + [{"id": "jun1", "name": "sj1",
                     "createdTime": "2026-06-20T00:00:00Z"},
                    {"id": "jun2", "name": "sj2",
                     "createdTime": "2026-06-10T00:00:00Z"},
                    {"id": "may1", "name": "sm1",
                     "createdTime": "2026-05-05T00:00:00Z"}])
        old_fl, old_fc = gd.files_list, gd.find_or_create_folder
        gd.files_list = lambda q, **kw: files
        gd.find_or_create_folder = lambda n, p: "SNAPS"
        old_keep = st.CFG.get("drive_table_snapshots")
        st.CFG["drive_table_snapshots"] = 3
        try:
            keep, drop = gd.snapshot_rotation_plan()
        finally:
            gd.files_list, gd.find_or_create_folder = old_fl, old_fc
            st.CFG["drive_table_snapshots"] = old_keep
        self.assertEqual([f["id"] for f in drop], ["jun2"])  # 2-я июньская
        self.assertIn("may1", [f["id"] for f in keep])       # по 1 на месяц


class TestNightlySchedule(unittest.TestCase):
    def test_due_at_hours(self):
        d = datetime.datetime(2026, 7, 9, 2, 59)
        self.assertFalse(ni.due_at(d, 3, None))
        d = datetime.datetime(2026, 7, 9, 3, 0)
        self.assertTrue(ni.due_at(d, 3, None))
        self.assertTrue(ni.due_at(d, 3, "2026-07-08"))
        self.assertFalse(ni.due_at(d, 3, "2026-07-09"))  # уже бегала сегодня

    def test_due_at_weekday_monthday(self):
        sun = datetime.datetime(2026, 7, 12, 5, 0)   # воскресенье
        self.assertTrue(ni.due_at(sun, 4, None, weekday=6))
        self.assertFalse(ni.due_at(sun, 4, None, weekday=0))
        first = datetime.datetime(2026, 8, 1, 5, 0)
        self.assertTrue(ni.due_at(first, 4, None, monthday=1))
        self.assertFalse(ni.due_at(sun, 4, None, monthday=1))

    def test_recon_text(self):
        live = {"10.20.50.1": "e0:7f:88:aa:bb:cc", "10.20.50.99": "AA:BB:CC:00:11:22"}
        cams = [{"ip": "10.20.50.1", "nmac": "e07f88aabbcc", "name": "AS-1"},
                {"ip": "10.20.50.2", "nmac": "", "name": "AS-2"}]
        text, stats = ni.recon_text(live, cams)
        self.assertIn("🆕1", stats)     # 10.20.50.99 не в инвентаре
        self.assertIn("❌1", stats)     # AS-2 не отвечает
        self.assertIn("10.20.50.99", text)
        self.assertIn("AS-2", text)


class TestFmtBuilders(unittest.TestCase):
    HDR = ["№", "Название", "IP-адрес", "MAC-адрес", "Статус монтажа",
           "Проверка (бот)"]

    def test_named_ranges(self):
        reqs = gf.named_range_reqs(set(), 7, self.HDR, 100)
        names = {r["addNamedRange"]["namedRange"]["name"] for r in reqs}
        self.assertEqual(names, {"IP_ADDR", "MAC_ADDR", "STATUS"})
        # уже есть — не дублируем
        reqs2 = gf.named_range_reqs({"IP_ADDR", "MAC_ADDR", "STATUS"},
                                    7, self.HDR, 100)
        self.assertEqual(reqs2, [])

    def test_freshness_rules(self):
        reqs = gf.freshness_reqs(7, 5, 100)
        self.assertEqual(len(reqs), 2)   # красная (7 дн.) + жёлтая (1 дн.)
        f0 = reqs[0]["addConditionalFormatRule"]["rule"]["booleanRule"][
            "condition"]["values"][0]["userEnteredValue"]
        self.assertIn("DATEVALUE(LEFT($F2,10))>7", f0)

    def test_filter_views(self):
        old = st.CFG.get("diff_subnets")
        st.CFG["diff_subnets"] = ["10.20.50"]
        try:
            reqs = gf.filter_view_reqs(set(), 7, self.HDR, 100, 6)
            titles = {r["addFilterView"]["filter"]["title"] for r in reqs}
            self.assertEqual(titles, {"Только офлайн", "Без MAC",
                                      "Подсеть 10.20.50.x"})
            # существующие пропускаются
            reqs2 = gf.filter_view_reqs(titles, 7, self.HDR, 100, 6)
            self.assertEqual(reqs2, [])
        finally:
            st.CFG["diff_subnets"] = old


class TestEventBuffer(unittest.TestCase):
    def test_note_event_buffers(self):
        with gs._ev_lock:
            gs._EV_BUF.clear()
        gs.note_event("test", "10.1.1.1", "AS-1", "деталь")
        with gs._ev_lock:
            self.assertEqual(len(gs._EV_BUF), 1)
            row = gs._EV_BUF[0]
            gs._EV_BUF.clear()
        self.assertEqual(row[1:4], ["test", "10.1.1.1", "AS-1"])

    def test_alert_hook_strips_html(self):
        with gs._ev_lock:
            gs._EV_BUF.clear()
        gs._alert_hook("🔴 <b>Камера упала</b>: 10.1.1.1\nвторая строка")
        with gs._ev_lock:
            row = gs._EV_BUF[0]
            gs._EV_BUF.clear()
        self.assertNotIn("<b>", row[4])
        self.assertNotIn("вторая", row[4])   # только первая строка


if __name__ == "__main__":
    unittest.main()
