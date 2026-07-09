# -*- coding: utf-8 -*-
"""Волна F: bot_dq (нормализация MAC/моделей, схема/lint, DQ-score, полнота),
bot_reports (парсер серийников), bot_backups (ротация, дифф двух xlsx),
bot_reconcile (журнал, CSV-срез, дайджест), миграция 259 — ВСЁ на временных
копиях; прод-файлы не читаются и не пишутся (кроме конфига bot_state)."""
import os
import sys
import json
import time
import shutil
import tempfile
import unittest
from unittest import mock

sys.path.insert(0, r"C:\Users\1\camera")

import bot_state as st  # noqa: E402
import bot_inventory as inv  # noqa: E402
import bot_dq as dq  # noqa: E402
import bot_reports as rp  # noqa: E402
import bot_backups as bk  # noqa: E402
import bot_reconcile as rc  # noqa: E402


def make_xlsx(path):
    """Мини-инвентарь: 4 листа с контролируемыми ошибками."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Все камеры"
    ws.append(["№", "Название (по ТЗ)", "IP-адрес", "MAC-адрес", "Расположение",
               "Объект", "Коммутатор", "IP коммутатора", "Порт", "VLAN",
               "MAC на порту", "Длина кабеля (м)", "Статус (2026-06-26)",
               "Модель камеры", "Статус (2026-07-08)"])
    ws.append([1, "AS-7C.01", "10.20.50.20", "E0:7F:88:06:43:51", "3 этаж",
               "Апартаменты", "SW-1", "10.10.60.59", "GE24", 1, 1, 84,
               "офлайн", "EVIDENCE Apix-Bullet/M4 27", "онлайн"])
    ws.append([2, "AS-7C.02", "10.20.50.21", "e0-7f-88-06-43-27", "3 этаж",
               "Апартаменты", "SW-1", "10.10.60.59", "GE19", 1, 1, 78,
               "онлайн", "Evidence APIX-Bullet/M4 27", "онлайн"])
    ws.append([3, None, "10.20.5.22", None, "3 этаж", "Апартаменты",
               "SW-2", "10.10.60.77", "порт9", 1, 1, None, None, None, None])
    ws.append([4, "AS-7C.04", "10.20.51.40", "AA:BB:CC:DD:EE:FF", "Подвалл",
               "Апартаменты", "SW-1", "10.10.60.59", "GE1", 1, 1, 10,
               None, "EVIDENCE Apix-Bullet/M4 27", "онлайн"])
    ws.append([None, None, None, None, "мусор в хвосте", None, None, None,
               None, None, None, None, None, None, None])
    ws2 = wb.create_sheet("Лист1")
    ws2.append(["Название коммутатора", "ip VLAN 1", "Серийник", "этаж"])
    ws2.append(["SW-1", "10.10.60.59", "NS220629012790001", "-1"])
    ws2.append(["SW-9", "10.10.60.99", "PS230101012780047", "2"])
    ws3 = wb.create_sheet("Неизвестные устройства")
    ws3.append(["MAC-адрес", "Вендор (OUI)", "Категория", "Коммутатор",
                "IP коммутатора", "Порт", "VLAN", "MAC на порту",
                "Видели на SW", "Примечание"])
    ws3.append(["E0:7F:88:06:43:51", "EVIDENCE", "Камера Apix", "SW-1",
                "10.10.60.59", "GE24", 1, 1, 1, None])
    ws3.append(["E0:7F:88:99:99:99", "EVIDENCE", "Камера Apix", "SW-1",
                "10.10.60.59", "GE2", 1, 1, 1, None])
    ws4 = wb.create_sheet("Изменённые")
    ws4.append(["№", "MAC-адрес", "Старый IP", "Новый IP", "Маска", "Шлюз",
                "Коммутатор", "IP коммутатора", "Порт", "Модель", "Серийник",
                "Прошивка", "Объект", "Что сделано", "Дата и время"])
    ws4.append([1, "E0:7F:88:03:BD:07", "192.168.0.250", "10.20.50.35",
                "255.255.255.0 (/24)", "10.20.50.254", "SW-8", "10.10.60.59",
                "GE10", "Evidence Apix-VDome/E5 271", "03BD07", "v5.1",
                "Апартаменты", "Заводская: смена IP", "2026-07-08 16:38"])
    ws4.append([2, "E0:7F:88:03:BD:07", "192.168.0.250", "10.20.50.99",
                "255.255.255.0 (/24)", "10.20.50.1", "SW-8", "10.10.60.59",
                "GE11", "Evidence Apix-VDome/E5 271", "03BD07", "v5.1",
                "Апартаменты", "дубль серийника", "08.07.2026"])
    wb.save(path)
    wb.close()


class Base(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tmp = tempfile.mkdtemp(prefix="wave_f_")
        cls.xlsx = os.path.join(cls.tmp, "inv.xlsx")
        make_xlsx(cls.xlsx)

    def setUp(self):
        self._back = {}
        over = {
            "inventory_xlsx": self.xlsx,
            "schema_path": os.path.join(self.tmp, "schema.json"),
            "dq_history_path": os.path.join(self.tmp, "dq_hist.json"),
            "models_path": os.path.join(self.tmp, "models.json"),
            "changelog_jsonl": os.path.join(self.tmp, "chg.jsonl"),
            "exports_dir": os.path.join(self.tmp, "exports"),
            "exports_git": False,
            "status_history_path": os.path.join(self.tmp, "sh.json"),
            "reconcile_state_path": os.path.join(self.tmp, "rs.json"),
            "facts_switches": os.path.join(self.tmp, "fsw.json"),
            "facts_cameras": os.path.join(self.tmp, "fc.json"),
            "loc_min_count": 2,
        }
        for k, v in over.items():
            self._back[k] = st.CFG.get(k)
            st.CFG[k] = v
        with inv._lock:
            inv._inv["mtime"] = None
            inv._unk["mtime"] = None
            inv._sw["mtime"] = None
            inv._fw["loaded"] = True
            inv._fw["data"] = {}

    def tearDown(self):
        for k, v in self._back.items():
            if v is None:
                st.CFG.pop(k, None)
            else:
                st.CFG[k] = v
        with inv._lock:
            inv._inv["mtime"] = None
            inv._unk["mtime"] = None
            inv._sw["mtime"] = None
            inv._fw["loaded"] = False
            inv._fw["data"] = {}


class TestMac(Base):
    def test_mac_canon(self):
        """253: канонизация разных написаний MAC."""
        self.assertEqual(dq.mac_canon("e0-7f-88.06:43 27"), "E0:7F:88:06:43:27")
        self.assertEqual(dq.mac_canon("E07F88064327"), "E0:7F:88:06:43:27")
        self.assertIsNone(dq.mac_canon("03BD07"))       # хвост — не MAC
        self.assertIsNone(dq.mac_canon(""))
        self.assertIsNone(dq.mac_canon(None))

    def test_mac_plan(self):
        """253: dry-run находит только неканоничные MAC."""
        plan = dq.mac_plan()
        self.assertEqual(len(plan), 1)
        sheet, rn, col, old, new = plan[0]
        self.assertEqual((sheet, rn, col), ("Все камеры", 3, "MAC-адрес"))
        self.assertEqual(new, "E0:7F:88:06:43:27")


class TestLint(Base):
    def test_lint_classes(self):
        """252/256/257/260/292/295: классы проблем находятся."""
        issues = dq.lint()
        codes = {i["code"] for i in issues}
        self.assertIn("256", codes)   # 10.20.5.22 вне политики + шлюз .1
        self.assertIn("253", codes)   # неканоничный MAC
        self.assertIn("292", codes)   # мусорная строка
        self.assertIn("295", codes)   # дата 08.07.2026
        self.assertIn("257", codes)   # дубль серийника 03BD07
        self.assertIn("260", codes)   # SW 10.10.60.77 нет в Лист1; SW-9 пустой
        ip_bad = [i for i in issues if i["code"] == "256"
                  and "10.20.5.22" in i["msg"]]
        self.assertTrue(ip_bad and ip_bad[0]["sev"] == "crit")

    def test_schema_created(self):
        """251: файл схемы создаётся при первом обращении."""
        dq.ensure_schema()
        self.assertTrue(os.path.exists(st.cget("schema_path")))
        data = json.load(open(st.cget("schema_path"), encoding="utf-8"))
        self.assertIn("Все камеры", data["sheets"])


class TestModels(Base):
    def test_model_map_variants(self):
        """254: канон — самое частое написание; разнобой найден."""
        mm = dq.model_map()
        canon = mm[dq._mkey("EVIDENCE Apix-Bullet/M4 27")]
        self.assertEqual(canon, "EVIDENCE Apix-Bullet/M4 27")  # 2 из 3
        var = dq.model_variants()
        self.assertIn(canon, var)
        self.assertIn("Evidence APIX-Bullet/M4 27", var[canon])


class TestScore(Base):
    def test_dq_score(self):
        """288: полная строка > дырявой; диапазон 0-100."""
        cams = {c.get("ip"): c for c in inv.cams()}
        full = dq.dq_score_row(cams["10.20.50.20"], facts_ok=False)
        holey = dq.dq_score_row(cams["10.20.5.22"], facts_ok=False)
        self.assertGreater(full, holey)
        self.assertGreaterEqual(holey, 0)
        self.assertLessEqual(full, 100)
        self.assertGreaterEqual(full, 80)

    def test_completeness_and_history(self):
        """287: полнота считается, снимок пишется в историю."""
        compl = dq.completeness()
        self.assertGreater(compl["Все камеры"]["IP-адрес"], 50)
        dq.note_dq(90.0, compl)
        h = dq.dq_history()
        self.assertEqual(len(h), 1)
        self.assertEqual(h[0]["score"], 90.0)
        dq.note_dq(91.0, compl)  # чаще раза в час — не пишем
        self.assertEqual(len(dq.dq_history()), 1)


class TestSerial(Base):
    def test_parse_serial_date(self):
        """263: NS/PS-YYMMDD -> дата; мусор -> None."""
        import datetime
        self.assertEqual(rp.parse_serial_date("NS220629012790001"),
                         datetime.date(2022, 6, 29))
        self.assertEqual(rp.parse_serial_date("ps230101012780047"),
                         datetime.date(2023, 1, 1))
        self.assertIsNone(rp.parse_serial_date("NS229999012790001"))
        self.assertIsNone(rp.parse_serial_date("03BD07"))
        self.assertIsNone(rp.parse_serial_date(None))

    def test_fleet_age(self):
        age = rp.fleet_age()
        self.assertEqual(len(age["switches"]), 2)
        self.assertIn(2022, age["by_year"])


class TestDiffXlsx(Base):
    def test_diff_two_copies(self):
        """297: дифф двух копий — added/removed/changed по UID."""
        a = os.path.join(self.tmp, "a.xlsx")
        b = os.path.join(self.tmp, "b.xlsx")
        shutil.copy2(self.xlsx, a)
        shutil.copy2(self.xlsx, b)
        import openpyxl
        wb = openpyxl.load_workbook(b)
        ws = wb["Все камеры"]
        ws.cell(row=2, column=5, value="7 этаж")       # location первой камеры
        ws.append([9, "AS-9X.01", "10.20.52.9", "E0:7F:88:AA:BB:CC", "лифт",
                   "Гостиница", "SW-9", "10.10.60.99", "GE5", 1, 1, 5,
                   None, "EVIDENCE Apix-Bullet/M4 27", "онлайн"])
        wb.save(a if False else b)
        wb.close()
        d = bk.diff_xlsx(a, b)
        self.assertEqual(d["added"], ["E0:7F:88:AA:BB:CC"])
        self.assertEqual(d["removed"], [])
        ch = [(u, f, va, vb) for u, f, va, vb in d["changed"]
              if f == "Расположение"]
        self.assertEqual(ch, [("E0:7F:88:06:43:51", "Расположение",
                               "3 этаж", "7 этаж")])


class TestRotation(Base):
    def test_rotation_plan(self):
        """269: дневные остаются, недельные схлопываются, старьё в drop."""
        now = time.time()
        baks = []
        for days in (1, 5, 20, 21, 22, 120):
            baks.append({"path": f"b{days}", "name": f"b{days}",
                         "mtime": now - days * 86400, "size": 1})
        keep, drop = bk.rotation_plan(baks)
        names_keep = {b["name"] for b in keep}
        self.assertIn("b1", names_keep)
        self.assertIn("b5", names_keep)
        self.assertIn("b120", {b["name"] for b in drop})   # старше 14+8*7
        # из 20/21/22 (могут попасть в 1-2 ISO-недели) хоть один в keep
        wk = [b for b in keep if b["name"] in ("b20", "b21", "b22")]
        self.assertTrue(wk)
        self.assertEqual(len(keep) + len(drop), len(baks))


class TestChangelog(Base):
    def test_record_and_history(self):
        """265: журнал пишется и читается; 268: дайджест видит правки."""
        rc.record_change("test", "Все камеры", "E0:7F:88:06:43:51",
                         "Примечание", "", "новая заметка")
        ev = rc.changelog_entries()
        self.assertEqual(len(ev), 1)
        self.assertEqual(ev[0]["field"], "Примечание")
        txt = rc.digest_text(7)
        self.assertIn("1 правок", txt)

    def test_csv_snapshot_deterministic(self):
        """267: два подряд среза байт-в-байт одинаковы."""
        p1 = rc.export_csv_snapshot()
        blobs1 = {p: open(p, "rb").read() for p in p1}
        p2 = rc.export_csv_snapshot()
        blobs2 = {p: open(p, "rb").read() for p in p2}
        self.assertEqual(blobs1, blobs2)
        self.assertEqual(len(p1), 4)


class TestApplyWrites(Base):
    def test_apply_writes_backup_and_log(self):
        """apply_writes: бэкап создаётся, ячейка пишется, журнал растёт."""
        with mock.patch.object(rc, "after_xlsx_write"):
            bak = dq.apply_writes(
                [("Все камеры", 3, "MAC-адрес", "e0-7f-88-06-43-27",
                  "E0:7F:88:06:43:27")], tag="test")
        self.assertTrue(os.path.exists(bak))
        self.assertIn(os.path.dirname(self.xlsx), bak)
        import openpyxl
        wb = openpyxl.load_workbook(self.xlsx)
        self.assertEqual(wb["Все камеры"].cell(row=3, column=4).value,
                         "E0:7F:88:06:43:27")
        wb.close()
        self.assertTrue(rc.changelog_entries())
        # вернуть как было (класс-фикстура общая)
        with mock.patch.object(rc, "after_xlsx_write"):
            dq.apply_writes([("Все камеры", 3, "MAC-адрес",
                              "E0:7F:88:06:43:27", "e0-7f-88-06-43-27")],
                            tag="test")


class TestMigrate(Base):
    def test_do_migrate_on_copy(self):
        """259: на КОПИИ статус-колонки схлопываются в Статус+Проверено."""
        cp = os.path.join(self.tmp, "mig.xlsx")
        shutil.copy2(self.xlsx, cp)
        import bot_dq_cmds as dqc
        r = dqc.do_migrate(cp)
        self.assertTrue(r["migrated"])
        self.assertEqual(r["dropped"], 1)
        self.assertEqual(r["checked"], "2026-07-08")
        self.assertGreaterEqual(r["events"], 2)  # онлайн/офлайн из 06-26
        import openpyxl
        wb = openpyxl.load_workbook(cp)
        hdr = [c.value for c in wb["Все камеры"][1]]
        wb.close()
        self.assertIn("Статус", hdr)
        self.assertIn("Проверено", hdr)
        self.assertNotIn("Статус (2026-06-26)", hdr)
        self.assertNotIn("Статус (2026-07-08)", hdr)


class TestSmis(Base):
    def test_smis_rows(self):
        """286: стабильная схема, uid — канонический MAC."""
        with mock.patch.object(dq, "_facts_port_ok", return_value=False):
            rows = rp.smis_rows()
        self.assertTrue(rows)
        self.assertEqual(set(rp.SMIS_FIELDS) - set(rows[0]), set())
        by_ip = {r["ip"]: r for r in rows}
        self.assertEqual(by_ip["10.20.50.21"]["uid"], "E0:7F:88:06:43:27")


class TestEnrichQueue(Base):
    def test_queue(self):
        """298: в очереди только камеры с пропусками."""
        import bot_enrich as en
        q = en.enrich_queue()
        self.assertEqual([c["ip"] for c in q], ["10.20.5.22"])


class TestFreshness(Base):
    def test_freshness_note(self):
        """299: нет файлов фактов -> предупреждение; свежие -> пусто."""
        self.assertIn("Факты устарели", rc.freshness_note())
        for k in ("facts_switches", "facts_cameras"):
            with open(st.cget(k), "w", encoding="utf-8") as f:
                f.write("[]")
        self.assertEqual(rc.freshness_note(), "")


if __name__ == "__main__":
    unittest.main(verbosity=2)
