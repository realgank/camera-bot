# -*- coding: utf-8 -*-
"""Волна H — транспорт и нормализация фактов коммутаторов (399).
Cross-24/FC (KT-NOS): web-API cgi/get.cgi|set.cgi, RSA-логин на чистом stdlib
(PKCS#1 v1.5 — только pow, без pycryptodome), кэш сессий с релогином.
Huawei VRP (S5736): plink subprocess, интерактивный shell через stdin
(exec-канал VRP не поддерживает), host-key TOFU: fingerprint из -batch
прогона кэшируется в _sw_hostkeys.json, перебор паролей #/@.
Парсеры Cross-24 (lang('…')-артефакты -> человеческие значения) и VRP-вывода —
чистые функции, тестируются на сохранённых образцах. Реестр свитчей =
_facts_switches.json + лист «Лист1» инвентаря + hw_switches конфига.
Запись (set.cgi / config-команды VRP) здесь НЕ выполняется — только транспорт;
save/reboot свитчей в этой волне не вызываются нигде."""
import os
import re
import time
import json
import base64
import secrets
import threading
import subprocess
from urllib.parse import quote

import requests

import bot_state as st
import bot_store as store
from bot_util import log, log_exc

_CF = getattr(subprocess, "CREATE_NO_WINDOW", 0)
HK_PATH = os.path.join(st.BASE, "_sw_hostkeys.json")


class SwError(Exception):
    """Ошибка транспорта/логина коммутатора."""


# ---------- RSA PKCS#1 v1.5 (encrypt-only, stdlib) ----------
def rsa_encrypt_b64(pwd: str, modulus_hex: str, exp_hex: str = "10001") -> str:
    """Шифрование пароля для home_loginAuth: PKCS#1 v1.5 + base64."""
    n = int(modulus_hex, 16)
    e = int(exp_hex, 16)
    k = (n.bit_length() + 7) // 8
    data = pwd.encode()
    if len(data) > k - 11:
        raise SwError("пароль длиннее, чем позволяет ключ")
    ps = b""
    while len(ps) < k - 3 - len(data):
        b = secrets.token_bytes(1)
        if b != b"\x00":
            ps += b
    em = b"\x00\x02" + ps + b"\x00" + data
    c = pow(int.from_bytes(em, "big"), e, n)
    return base64.b64encode(c.to_bytes(k, "big")).decode()


# ---------- Cross-24 web-API (кэш сессий) ----------
_sessions: dict = {}          # ip -> requests.Session (залогиненная)
_sess_lock = threading.RLock()


def _tmo() -> float:
    return float(st.cget("sw_http_timeout_s"))


def _ts() -> int:
    return int(time.time() * 1000)


def _login(ip: str) -> requests.Session:
    s = requests.Session()
    base = f"http://{ip}"
    r = s.get(f"{base}/cgi/get.cgi?cmd=home_login&dummy={_ts()}", timeout=_tmo())
    try:
        modulus = r.json()["data"]["modulus"]
    except Exception:
        raise SwError(f"{ip}: home_login не отдал modulus")
    enc = rsa_encrypt_b64(str(st.cget("sw_pass")), modulus)
    body = (f"_ds=1&username={quote(str(st.cget('sw_user')), safe='')}"
            f"&password={quote(enc, safe='')}&_de=1")
    s.post(f"{base}/cgi/set.cgi?cmd=home_loginAuth&dummy={_ts()}", data=body,
           headers={"Content-Type": "application/x-www-form-urlencoded"},
           timeout=_tmo())
    for _ in range(15):
        time.sleep(0.4)
        try:
            j = s.get(f"{base}/cgi/get.cgi?cmd=home_loginStatus&dummy={_ts()}",
                      timeout=_tmo()).json()
        except Exception:
            continue
        status = (j.get("data") or {}).get("status")
        if status == "ok":
            return s
        if status == "fail":
            raise SwError(f"{ip}: логин web-API отвергнут "
                          f"({(j.get('data') or {}).get('failReason', '')})")
    raise SwError(f"{ip}: логин web-API не подтвердился (таймаут)")


def cross24_reset(ip: str) -> None:
    with _sess_lock:
        s = _sessions.pop(ip, None)
    if s:
        try:
            s.close()
        except Exception:
            pass


def _c24_call(ip: str, kind: str, cmd: str, pairs=None) -> dict:
    with _sess_lock:
        s = _sessions.get(ip)
    if s is None:
        s = _login(ip)
        with _sess_lock:
            _sessions[ip] = s
    base = f"http://{ip}"
    if kind == "get":
        r = s.get(f"{base}/cgi/get.cgi?cmd={cmd}&dummy={_ts()}", timeout=_tmo())
    else:
        if isinstance(pairs, dict):
            pairs = list(pairs.items())
        body = "&".join(f"{k}={quote(str(v), safe='')}" for k, v in pairs or [])
        r = s.post(f"{base}/cgi/set.cgi?cmd={cmd}&dummy={_ts()}",
                   data=f"_ds=1&{body}&_de=1", timeout=_tmo(),
                   headers={"Content-Type": "application/x-www-form-urlencoded"})
    try:
        return r.json()
    except Exception:
        raise SwError(f"{ip}: {cmd} вернул не-JSON (HTTP {r.status_code})")


def _c24(ip: str, kind: str, cmd: str, pairs=None) -> dict:
    """Вызов с одним релогином при протухшей сессии."""
    try:
        j = _c24_call(ip, kind, cmd, pairs)
    except (requests.RequestException, SwError):
        cross24_reset(ip)
        j = _c24_call(ip, kind, cmd, pairs)
    msg = str(j.get("msgType") or "")
    if "Login" in msg or "login" in msg:      # errLoginUserInvalid и т.п.
        cross24_reset(ip)
        j = _c24_call(ip, kind, cmd, pairs)
    return j


def cross24_get(ip: str, cmd: str) -> dict:
    """GET get.cgi?cmd=…; возвращает data-словарь."""
    j = _c24(ip, "get", cmd)
    if "data" not in j:
        raise SwError(f"{ip}: {cmd}: нет data в ответе ({str(j)[:120]})")
    return j["data"]


def cross24_set(ip: str, cmd: str, pairs) -> dict:
    """SET set.cgi?cmd=… (обёртка _ds/_de). Ответ целиком."""
    return _c24(ip, "set", cmd, pairs)


# ---------- парсеры Cross-24 (lang-артефакты -> значения), 399 ----------
_LANG_ARGS = re.compile(r"lang\('[^']*','[^']*',\[([\d,\s-]*)\]\)")
_LANG_LBL = re.compile(r"lang\('[^']*','(?:lbl|txt)?([^']*)'")


def lang_args(s: str) -> list:
    """'lang('sys','txtSysUptimeArg',[63,0,27,48])' -> [63, 0, 27, 48]."""
    m = _LANG_ARGS.search(str(s or ""))
    if not m:
        return []
    return [int(x) for x in m.group(1).split(",") if x.strip().lstrip("-").isdigit()]


def lang_label(s: str) -> str:
    """'lang('lldp','lblMacAddr')' -> 'MacAddr'; обычная строка — как есть."""
    s = str(s or "")
    m = _LANG_LBL.search(s)
    return m.group(1) if m else s


def c24_uptime_s(sys_d: dict):
    """sysUpTime lang-args [дни, часы, мин, сек] -> секунды (или None)."""
    a = lang_args((sys_d or {}).get("sysUpTime"))
    if len(a) != 4:
        return None
    d, h, m, s = a
    return ((d * 24 + h) * 60 + m) * 60 + s


def c24_methods(sys_d: dict) -> dict:
    """methods -> {'telnet': bool, 'ssh': …, 'http', 'https', 'snmp'}."""
    out = {}
    for m in (sys_d or {}).get("methods") or []:
        name = lang_label(m.get("txt")).lower()
        out[name] = bool(m.get("state"))
    return out


def port_index(name: str):
    """'GE7' -> 6, 'TE2' -> 25 (24 медных + 4 uplink); None — не порт."""
    m = re.fullmatch(r"(GE|TE)(\d+)", str(name or "").strip(), re.I)
    if not m:
        return None
    n = int(m.group(2))
    return (n - 1) if m.group(1).upper() == "GE" else (24 + n - 1)


def norm_switch(e: dict) -> dict:
    """399: сырая запись _facts_switches.json -> единая нормализованная схема."""
    sysd = e.get("sys") or {}
    lldp = []
    for n in e.get("lldp") or []:
        lldp.append({"port": n.get("localPort"),
                     "id": (n.get("chassisId") or "").upper(),
                     "port_id": n.get("portId") or "",
                     "name": n.get("sysName") or ""})
    return {"ip": e.get("ip"), "ok": bool(e.get("ok")), "err": e.get("err"),
            "kind": "cross24",
            "host": sysd.get("hostname") or e.get("ip"),
            "sn": sysd.get("syssn"), "mac": (sysd.get("sysMac") or "").upper(),
            "fw": sysd.get("fwVer"), "loader": sysd.get("loaderVer"),
            "uptime_s": c24_uptime_s(sysd),
            "time_epoch": sysd.get("sec"),
            "time_str": (sysd.get("sysCurrTime") or "").strip(),
            "location": sysd.get("location"), "contact": sysd.get("contact"),
            "methods": c24_methods(sysd),
            "lldp": lldp, "macs_n": len(e.get("mac_table") or []),
            "mac_table": e.get("mac_table") or []}


# ---------- факты (кэш по mtime) ----------
_facts = {"mtime": None, "raw": [], "norm": []}
_flock = threading.RLock()


def facts() -> list:
    """Нормализованные записи фактов (перечитка по mtime)."""
    path = st.cget("facts_switches")
    with _flock:
        try:
            mt = os.path.getmtime(path)
        except OSError:
            return list(_facts["norm"])
        if _facts["mtime"] != mt:
            try:
                with open(path, encoding="utf-8") as f:
                    raw = json.load(f)
                _facts.update(mtime=mt, raw=raw,
                              norm=[norm_switch(e) for e in raw])
                log(f"sw_api: перечитаны факты свитчей ({len(raw)} записей)")
            except Exception:
                log_exc("sw_api: не смог прочитать _facts_switches.json")
        return list(_facts["norm"])


def facts_age_days():
    try:
        return (time.time() - os.path.getmtime(st.cget("facts_switches"))) / 86400
    except OSError:
        return None


def age_note() -> str:
    """Штамп возраста фактов для /topo и пр. (397)."""
    d = facts_age_days()
    if d is None:
        return "факты свитчей отсутствуют"
    return f"факты от {time.strftime('%d.%m', time.localtime(time.time() - d * 86400))} ({d:.0f} дн. назад)"


def by_ip(ip: str):
    for e in facts():
        if e["ip"] == ip:
            return e
    return None


def switch_macs() -> dict:
    """sysMac -> ip всех известных Cross-24 (для распознавания LLDP-соседей)."""
    return {e["mac"]: e["ip"] for e in facts() if e.get("mac")}


# ---------- реестр свитчей: факты + «Лист1» + hw_switches ----------
_sheet = {"mtime": None, "rows": []}


def sheet_switches() -> list:
    """Лист «Лист1» инвентаря: [{name, room, floor, ip, sn}]."""
    path = st.cget("inventory_xlsx")
    with _flock:
        try:
            mt = os.path.getmtime(path)
        except OSError:
            return list(_sheet["rows"])
        if _sheet["mtime"] == mt:
            return list(_sheet["rows"])
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            rows = []
            if "Лист1" in wb.sheetnames:
                it = wb["Лист1"].iter_rows(values_only=True)
                hdr = [str(h or "") for h in next(it, [])]
                def col(sub):
                    for i, h in enumerate(hdr):
                        if sub.lower() in h.lower():
                            return i
                    return None
                ci = {"name": col("Название"), "room": col("помещение"),
                      "floor": col("этаж"), "ip": col("ip"), "sn": col("Серийник")}
                for r in it:
                    def g(k):
                        i = ci.get(k)
                        return r[i] if i is not None and i < len(r) else None
                    if g("ip"):
                        rows.append({"name": g("name"), "room": g("room"),
                                     "floor": g("floor"),
                                     "ip": str(g("ip")).strip(), "sn": g("sn")})
            wb.close()
            _sheet.update(mtime=mt, rows=rows)
        except Exception:
            log_exc("sw_api: не смог прочитать «Лист1»")
        return list(_sheet["rows"])


def registry() -> list:
    """Единый реестр: [{ip, host, kind, room, floor, sn, ok}] — cross24 из
    фактов и «Лист1», huawei из конфига hw_switches."""
    reg: dict = {}
    for row in sheet_switches():
        reg[row["ip"]] = {"ip": row["ip"], "host": row.get("name"),
                          "kind": "cross24", "room": row.get("room"),
                          "floor": row.get("floor"), "sn": row.get("sn"),
                          "ok": None}
    for e in facts():
        r = reg.setdefault(e["ip"], {"ip": e["ip"], "kind": "cross24",
                                     "room": None, "floor": None, "sn": None})
        r.update(host=e["host"], sn=e.get("sn") or r.get("sn"), ok=e["ok"],
                 fw=e.get("fw"), uptime_s=e.get("uptime_s"))
    for ip in st.cget("hw_switches") or []:
        reg.setdefault(str(ip), {"ip": str(ip), "host": None, "room": None,
                                 "floor": None, "sn": None, "ok": None}) \
           .update(kind="huawei")
    return sorted(reg.values(),
                  key=lambda r: tuple(int(o) for o in r["ip"].split(".")))


def find_switch(arg: str):
    """IP или имя (fuzzy) -> запись реестра или None."""
    a = (arg or "").strip()
    if not a:
        return None
    for r in registry():
        if r["ip"] == a:
            return r
    low = a.lower().replace(" ", "")
    for r in registry():
        if (str(r.get("host") or "").lower().replace(" ", "") == low
                or low in str(r.get("host") or "").lower()):
            return r
    return None


# Huawei-транспорт и парсеры VRP вынесены в bot_sw_hw (лимит 500 строк)
from bot_sw_hw import (plink_ok, huawei_cli, hw_section, hw_sysname,  # noqa: F401,E402
    hw_short_if, hw_parse_version, hw_parse_poe_ports, hw_parse_poe_info,
    hw_parse_int_brief, hw_parse_lldp_brief, hw_parse_temperature,
    hw_parse_ntp, hw_parse_clock, hw_parse_mac_table)


# ---------- двухшаговое подтверждение с TTL (общее для волны H) ----------
class Confirm:
    """Хранилище отложенных операций: put() на шаге dry-run, take() на
    финальной кнопке (None — протухло по TTL или не было)."""

    def __init__(self):
        self._p: dict = {}
        self._lock = threading.Lock()

    def put(self, key: str, payload) -> None:
        with self._lock:
            self._p[key] = (time.time(), payload)

    def take(self, key: str):
        with self._lock:
            v = self._p.pop(key, None)
        if not v:
            return None
        ts, payload = v
        if time.time() - ts > float(st.cget("sw_confirm_ttl_s")):
            return None
        return payload


HANDLERS: dict = {}
ALIASES: dict = {}
CALLBACKS: dict = {}
