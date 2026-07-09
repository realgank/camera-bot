# -*- coding: utf-8 -*-
"""Инвентарь камер: кэш Все_камеры.xlsx (перечитка по mtime), поиск по
IP/имени/MAC/серийнику с fuzzy-нормализацией (as7c01 -> AS-7C.01), карточка
камеры, порт свитча из _facts_switches.json, кэш _facts_cameras.json,
накопительный кэш прошивок (ONVIF), запись /note с автобэкапом.
Волна B: I1, I2, I3, I8, I9, I42, I43, часть I31/I47/I48.
Запись в xlsx — ТОЛЬКО write_note() и только после автобэкапа (I42)."""
import os
import re
import json
import shutil
import difflib
import datetime
import threading

import bot_state as st
from bot_util import log, log_exc

SHEET_MAIN = "Все камеры"
FW_PATH = os.path.join(st.BASE, "_fw_cache.json")

# поле карточки -> заголовок колонки в листе «Все камеры»
COLS = {"n": "№", "name": "Название (по ТЗ)", "ip": "IP-адрес", "mac": "MAC-адрес",
        "location": "Расположение", "obj": "Объект", "switch": "Коммутатор",
        "sw_ip": "IP коммутатора", "port": "Порт", "vlan": "VLAN",
        "cable": "Длина кабеля (м)", "model": "Модель камеры", "note": "Примечание"}

_lock = threading.RLock()
_inv = {"mtime": None, "cams": [], "by_ip": {}, "status_hdr": ""}
_sw = {"mtime": None, "by_mac": {}, "density": {}}
_fc = {"mtime": None, "by_ip": {}}
_snap = {"mtime": None, "by_ip": {}}
_fw = {"loaded": False, "data": {}}


def inv_path():
    return st.cget("inventory_xlsx")


# ---------- I8: fuzzy-нормализация ----------
def norm_name(s):
    """'AS-7C.01' -> 'as7c01' — сравнение имён без регистра и разделителей."""
    return re.sub(r"[^a-zа-яё0-9]", "", str(s or "").lower())


def norm_mac(s):
    """MAC/серийник -> hex-строка без разделителей в нижнем регистре."""
    return re.sub(r"[^0-9a-f]", "", str(s or "").lower())


# ---------- I1: кэш инвентаря по mtime ----------
def _load_inv():
    import openpyxl
    path = inv_path()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[SHEET_MAIN]
    it = ws.iter_rows(values_only=True)
    headers = list(next(it))
    idx = {str(h): i for i, h in enumerate(headers) if h}
    col = {k: idx.get(v) for k, v in COLS.items()}
    stat_cols = [i for i, h in enumerate(headers) if h and str(h).startswith("Статус")]
    scol = stat_cols[-1] if stat_cols else None
    cams, by_ip = [], {}
    for rn, row in enumerate(it, start=2):
        rec = {}
        for k, c in col.items():
            rec[k] = row[c] if c is not None and c < len(row) else None
        if not rec.get("ip") and not rec.get("mac") and not rec.get("name"):
            continue
        rec["ip"] = str(rec["ip"]).strip() if rec.get("ip") else None
        rec["row"] = rn
        rec["status"] = row[scol] if scol is not None and scol < len(row) else None
        rec["nname"] = norm_name(rec.get("name"))
        rec["nmac"] = norm_mac(rec.get("mac"))
        cams.append(rec)
        if rec["ip"] and rec["ip"] not in by_ip:
            by_ip[rec["ip"]] = rec
    wb.close()
    return {"mtime": os.path.getmtime(path), "cams": cams, "by_ip": by_ip,
            "status_hdr": str(headers[scol]) if scol is not None else "Статус"}


def _fresh(cache, path, loader, what):
    """Общий паттерн «перечитать по mtime»."""
    try:
        mt = os.path.getmtime(path)
    except OSError:
        return cache
    if cache.get("mtime") != mt:
        try:
            cache.update(loader())
            log(f"inventory: перечитан {what} (mtime {mt:.0f})")
        except Exception:
            log_exc(f"inventory: не смог прочитать {what}")
    return cache


def cams():
    with _lock:
        return _fresh(_inv, inv_path(), _load_inv, "Все_камеры.xlsx")["cams"]


def get(ip):
    """Запись инвентаря по IP или None."""
    with _lock:
        return _fresh(_inv, inv_path(), _load_inv, "Все_камеры.xlsx")["by_ip"].get(ip)


def status_header():
    with _lock:
        _fresh(_inv, inv_path(), _load_inv, "Все_камеры.xlsx")
        return _inv["status_hdr"]


# ---------- I2: поиск по IP/имени/MAC/серийнику ----------
def search(q):
    """Список записей: IP -> точное; hex 12 -> MAC; hex 4-11 -> хвост MAC
    (серийник Apix = хвост MAC); иначе — по имени/расположению (fuzzy)."""
    q = (q or "").strip()
    if not q:
        return []
    all_cams = cams()
    if re.fullmatch(r"\d{1,3}(\.\d{1,3}){3}", q):
        rec = get(q)
        return [rec] if rec else []
    h = norm_mac(q)
    if h and re.fullmatch(r"[0-9a-f]+", q.lower().replace(":", "").replace("-", "")
                          .replace(".", "")) and len(h) >= 4:
        if len(h) == 12:
            res = [c for c in all_cams if c["nmac"] == h]
        else:
            res = [c for c in all_cams if c["nmac"].endswith(h)]
        if res:
            return res
    nn = norm_name(q)
    if not nn:
        return []
    exact = [c for c in all_cams if c["nname"] == nn]
    if exact:
        return exact
    starts = [c for c in all_cams if c["nname"].startswith(nn)]
    contains = [c for c in all_cams if nn in c["nname"] and c not in starts]
    loc = [c for c in all_cams if nn in norm_name(c.get("location"))
           and c not in starts and c not in contains]
    return starts + contains + loc


def suggest(q, n=3):
    """Похожие имена для «не найдено» (difflib по нормализованным именам)."""
    nn = norm_name(q)
    names = {c["nname"]: c["name"] for c in cams() if c["nname"]}
    return [names[m] for m in difflib.get_close_matches(nn, names, n=n, cutoff=0.55)]


def resolve_ip(q):
    """Имя/MAC/серийник -> IP (если нашлась ровно одна камера с IP)."""
    res = [c for c in search(q) if c.get("ip")]
    return res[0]["ip"] if len(res) == 1 else None


def label(ip):
    """Короткая подпись для /find, /shot: 'AS-7C.01 · Лестница 5' или None."""
    rec = get(ip)
    if not rec:
        return None
    parts = [str(rec[k]) for k in ("name", "location") if rec.get(k)]
    return " · ".join(parts) if parts else None


# ---------- I3: карточка камеры ----------
def card_text(rec, live_mac=None):
    from bot_util import esc
    name = rec.get("name") or "(без имени в инвентаре)"
    lines = [f"📷 <b>{esc(name)}</b> — <code>{esc(rec.get('ip') or '?')}</code> "
             f"(№{esc(rec.get('n') or '?')})"]
    if rec.get("model"):
        lines.append(f"🎥 {esc(rec['model'])}")
    mac = rec.get("mac")
    if mac:
        m = f"MAC: <code>{esc(mac)}</code>"
        if live_mac and norm_mac(live_mac) != rec["nmac"]:
            m += f" ⚠️ сейчас в сети {esc(live_mac)} — MAC сменился!"
        lines.append(m)
    locp = " · ".join(str(rec[k]) for k in ("location", "obj") if rec.get(k))
    if locp:
        lines.append(f"📍 {esc(locp)}")
    swp = []
    if rec.get("switch"):
        swp.append(str(rec["switch"]))
    if rec.get("sw_ip"):
        swp.append(f"<code>{esc(rec['sw_ip'])}</code>")
    if rec.get("port"):
        swp.append(f"порт {esc(rec['port'])}")
    if rec.get("vlan") not in (None, ""):
        swp.append(f"VLAN {esc(rec['vlan'])}")
    if swp:
        lines.append("🔌 " + " · ".join(swp))
    if rec.get("cable"):
        lines.append(f"📏 кабель ~{esc(rec['cable'])} м")
    if rec.get("status"):
        lines.append(f"📶 {esc(status_header())}: {esc(rec['status'])}")
    fw = fw_cache().get(rec.get("ip") or "")
    if fw:
        lines.append(f"🧩 fw {esc(fw.get('fw'))} · sn {esc(fw.get('serial'))} "
                     f"(ONVIF {esc(fw.get('ts', '?')[:16])})")
    if rec.get("note"):
        lines.append(f"📝 {esc(rec['note'])}")
    dl = drive_link(rec.get("ip"))
    if dl:  # I31: индекс снимков Drive
        lines.append(f'🖼 <a href="{dl}">снимок в Drive</a>')
    return "\n".join(lines)


def ip_card_text(ip):
    """Карточка по голому IP: инвентарная, либо «нет в инвентаре»."""
    rec = get(ip)
    if rec:
        return card_text(rec)
    return (f"<code>{ip}</code> — 🆕 в инвентаре НЕ найдена.\n"
            f"Что сделать с этим адресом?")


# ---------- I9: порт свитча из _facts_switches.json ----------
def _load_sw():
    with open(st.cget("facts_switches"), encoding="utf-8") as f:
        data = json.load(f)
    by_mac, density = {}, {}
    for e in data:
        if not e.get("ok"):
            continue
        host = (e.get("sys") or {}).get("hostname") or e.get("ip")
        for m in e.get("mac_table") or []:
            key = (e["ip"], m.get("port"))
            density[key] = density.get(key, 0) + 1
            nm = norm_mac(m.get("mac"))
            if nm:
                by_mac.setdefault(nm, []).append(
                    {"host": host, "sw_ip": e["ip"], "port": m.get("port"),
                     "vlan": m.get("vlan")})
    return {"mtime": os.path.getmtime(st.cget("facts_switches")),
            "by_mac": by_mac, "density": density}


def switch_ports(mac):
    """Где видели MAC: [{host, sw_ip, port, vlan, density}], access-порты первыми."""
    with _lock:
        _fresh(_sw, st.cget("facts_switches"), _load_sw, "_facts_switches.json")
        hits = list(_sw["by_mac"].get(norm_mac(mac), []))
        for h in hits:
            h["density"] = _sw["density"].get((h["sw_ip"], h["port"]), 0)
    return sorted(hits, key=lambda x: x["density"])


# ---------- кэш _facts_cameras.json (I47) ----------
def _load_fc():
    with open(st.cget("facts_cameras"), encoding="utf-8") as f:
        data = json.load(f)
    return {"mtime": os.path.getmtime(st.cget("facts_cameras")),
            "by_ip": {e["ip"]: e for e in data if e.get("ip")}}


def facts_cam(ip):
    with _lock:
        _fresh(_fc, st.cget("facts_cameras"), _load_fc, "_facts_cameras.json")
        return _fc["by_ip"].get(ip)


# ---------- I31: индекс снимков Drive (пишет bot_sheets) ----------
def drive_link(ip):
    if not ip:
        return None
    path = st.cget("snap_index_path")
    with _lock:
        _fresh(_snap, path, lambda: {
            "mtime": os.path.getmtime(path),
            "by_ip": json.load(open(path, encoding="utf-8"))}, "индекс снимков")
        e = _snap["by_ip"].get(ip)
    return f"https://drive.google.com/file/d/{e['id']}/view" if e and e.get("id") else None


# ---------- I48: накопительный кэш прошивок ----------
def fw_cache():
    with _lock:
        if not _fw["loaded"]:
            try:
                with open(FW_PATH, encoding="utf-8") as f:
                    _fw["data"] = json.load(f)
            except Exception:
                _fw["data"] = {}
            _fw["loaded"] = True
        return dict(_fw["data"])


def note_onvif(ip, info):
    """Успешный device_info -> копим модель/прошивку/серийник (для /fw, /verify)."""
    if not info or not info.get("model"):
        return
    rec = {"model": info.get("model"), "fw": info.get("firmware"),
           "serial": info.get("serial"),
           "manufacturer": info.get("manufacturer"),
           "ts": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
    try:
        with _lock:
            fw_cache()
            _fw["data"][ip] = rec
            tmp = FW_PATH + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump(_fw["data"], f, ensure_ascii=False, indent=1)
            os.replace(tmp, FW_PATH)
    except Exception:
        log_exc("note_onvif: не смог сохранить _fw_cache.json")


# ---------- I35: лист «Неизвестные устройства» (read-only) ----------
UNK_SHEET = "Неизвестные устройства"
_unk = {"mtime": None, "hdr": [], "rows": []}


def _load_unk():
    import openpyxl
    path = inv_path()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if UNK_SHEET not in wb.sheetnames:
            return {"mtime": os.path.getmtime(path), "hdr": [], "rows": []}
        it = wb[UNK_SHEET].iter_rows(values_only=True)
        hdr = [str(h or "") for h in next(it, [])]
        rows = [list(r) for r in it if any(c not in (None, "") for c in r)]
        return {"mtime": os.path.getmtime(path), "hdr": hdr, "rows": rows}
    finally:
        wb.close()


def unknown_devices():
    """(заголовки, строки) листа «Неизвестные устройства» — только чтение."""
    with _lock:
        _fresh(_unk, inv_path(), _load_unk, "лист «Неизвестные устройства»")
        return list(_unk["hdr"]), list(_unk["rows"])


# ---------- I42 + I43: запись примечания с автобэкапом ----------
def write_note(ip, text):
    """Пишет текст в колонку «Примечание» строки камеры (создаёт колонку при
    отсутствии). ПЕРЕД записью — обязательный бэкап *_bot.xlsx (I42).
    Возвращает (путь_бэкапа, row) или бросает исключение."""
    import openpyxl
    path = inv_path()
    rec = get(ip)
    if not rec:
        raise ValueError(f"IP {ip} не найден в инвентаре")
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    bak = os.path.join(os.path.dirname(path), f"Все_камеры.backup.{ts}_bot.xlsx")
    shutil.copy2(path, bak)  # I42
    wb = openpyxl.load_workbook(path)
    try:
        ws = wb[SHEET_MAIN]
        headers = [c.value for c in ws[1]]
        if "Примечание" in headers:
            nc = headers.index("Примечание") + 1
        else:
            nc = len(headers) + 1
            ws.cell(row=1, column=nc, value="Примечание")
        ws.cell(row=rec["row"], column=nc, value=text)
        wb.save(path)
    finally:
        wb.close()
    with _lock:
        _inv["mtime"] = None  # перечитать при следующем запросе
    try:  # Волна F (265/267): журнал изменений + CSV-срез (best-effort)
        import bot_reconcile
        bot_reconcile.record_change("note", SHEET_MAIN, rec.get("mac") or ip,
                                    "Примечание", rec.get("note"), text)
        bot_reconcile.after_xlsx_write(f"/note {ip}")
    except Exception:
        log_exc("note: журнал изменений не записался")
    log(f"note: ip={ip} row={rec['row']} бэкап={os.path.basename(bak)}")
    return bak, rec["row"]
