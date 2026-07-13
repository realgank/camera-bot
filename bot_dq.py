# -*- coding: utf-8 -*-
"""Волна F — движок качества данных инвентаря (без Telegram-зависимостей):
251 схема inventory_schema.json, 252 lint-проверки, 253 план нормализации MAC,
254 канонизация моделей, 255 словарь локаций, 256 IP-политика, 257 дубли
серийников, 258 OUI-сверка, 260 ссылочная целостность, 287 полнота с трендом
(_dq_history.json), 288 DQ-score, 292 мусорные строки, 293 VLAN-правила,
295 форматы дат, 300 сводка. Плюс общий писатель в xlsx с автобэкапом
(backup_xlsx/apply_writes) для 253/273/298/259. Только stdlib+openpyxl."""
import os
import re
import json
import time
import shutil
import datetime
import collections

import bot_state as st
import bot_inventory as inv
import bot_store as store
from bot_util import log, log_exc

MAC_RE = re.compile(r"^[0-9A-F]{2}(:[0-9A-F]{2}){5}$")
IP_RE = re.compile(r"^\d{1,3}(\.\d{1,3}){3}$")
PORT_RE = re.compile(r"^(GE|TE|XGE|GI)\s?\d{1,2}$", re.IGNORECASE)
DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}([ T]\d{2}:\d{2}(:\d{2})?)?$")
HDR_DATE_RE = re.compile(r"\((\d{4}-\d{2}-\d{2})\)")

# ---------- 251: схема инвентаря ----------
DEFAULT_SCHEMA = {
    "version": 1,
    "sheets": {
        "Все камеры": {
            "key": "MAC-адрес",
            "required": ["IP-адрес", "MAC-адрес"],
            "columns": {
                "№": {"type": "int"}, "Название (по ТЗ)": {"type": "str"},
                "IP-адрес": {"type": "ip", "policy": "cam"},
                "MAC-адрес": {"type": "mac"},
                "Расположение": {"type": "loc"}, "Объект": {"type": "loc"},
                "Коммутатор": {"type": "str"},
                "IP коммутатора": {"type": "ip", "policy": "sw"},
                "Порт": {"type": "port"}, "VLAN": {"type": "int"},
                "MAC на порту": {"type": "int"},
                "Длина кабеля (м)": {"type": "float"},
                "Модель камеры": {"type": "model"},
                "Примечание": {"type": "str"}},
            "extra_prefixes": ["Статус", "Снимок", "Online", "Проверка",
                               "Проверено"]},
        "Лист1": {
            "key": "ip VLAN 1", "required": ["Название коммутатора"],
            "columns": {
                "Название коммутатора": {"type": "str"},
                "ip VLAN 1": {"type": "ip", "policy": "sw"},
                "Серийник": {"type": "serial"}, "этаж": {"type": "str"}},
            "extra_prefixes": [""]},
        "Неизвестные устройства": {
            "key": "MAC-адрес", "required": ["MAC-адрес"],
            "columns": {
                "MAC-адрес": {"type": "mac"},
                "IP коммутатора": {"type": "ip", "policy": "sw"},
                "Порт": {"type": "port"}, "VLAN": {"type": "int"}},
            "extra_prefixes": [""]},
        "Изменённые": {
            "key": "MAC-адрес", "required": ["MAC-адрес", "Новый IP"],
            "columns": {
                "№": {"type": "int"}, "MAC-адрес": {"type": "mac"},
                "Старый IP": {"type": "ip", "policy": "any"},
                "Новый IP": {"type": "ip", "policy": "cam"},
                "Шлюз": {"type": "gw"}, "Порт": {"type": "port"},
                "IP коммутатора": {"type": "ip", "policy": "sw"},
                "Модель": {"type": "model"}, "Серийник": {"type": "serial"},
                "Дата и время": {"type": "dt"}},
            "extra_prefixes": [""]},
    },
}


def ensure_schema() -> dict:
    """251: создаёт inventory_schema.json при отсутствии, возвращает схему."""
    path = st.cget("schema_path")
    data = store.jload(path, {})
    if data.get("sheets"):
        return data
    try:
        store.jsave(path, DEFAULT_SCHEMA)
        log(f"dq: создана схема {os.path.basename(path)}")
    except Exception:
        log_exc("dq: не смог записать схему")
    return json.loads(json.dumps(DEFAULT_SCHEMA))


# ---------- чтение всех листов ----------
def read_all(path=None) -> dict:
    """{title: {'hdr': [...], 'rows': [[...], ...]}} — строки с Excel-строки 2."""
    import openpyxl
    path = path or inv.inv_path()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {}
    try:
        for ws in wb.worksheets:
            it = ws.iter_rows(values_only=True)
            hdr = [str(h) if h is not None else "" for h in next(it, [])]
            rows = [list(r) for r in it]
            while rows and all(c in (None, "") for c in rows[-1]):
                rows.pop()
            out[ws.title] = {"hdr": hdr, "rows": rows}
    finally:
        wb.close()
    return out


def _cell(row, hdr, name):
    try:
        i = hdr.index(name)
    except ValueError:
        return None
    return row[i] if i < len(row) else None


# ---------- 253: нормализация MAC ----------
def mac_canon(s):
    """'e0-7f-88.064351' -> 'E0:7F:88:06:43:51'; None — не полный MAC."""
    h = re.sub(r"[^0-9A-Fa-f]", "", str(s or ""))
    if len(h) != 12:
        return None
    h = h.upper()
    return ":".join(h[i:i + 2] for i in range(0, 12, 2))


def mac_plan(data=None, schema=None) -> list:
    """253: [(лист, excel_row, колонка, старое, новое)] — что нормализуется."""
    data = data or read_all()
    schema = schema or ensure_schema()
    plan = []
    for sheet, spec in schema["sheets"].items():
        d = data.get(sheet)
        if not d:
            continue
        cols = [c for c, m in spec["columns"].items() if m.get("type") == "mac"]
        for rn, row in enumerate(d["rows"], start=2):
            for c in cols:
                v = _cell(row, d["hdr"], c)
                if v in (None, ""):
                    continue
                canon = mac_canon(v)
                if canon and str(v) != canon:
                    plan.append((sheet, rn, c, str(v), canon))
    return plan


# ---------- 254: канонизация моделей ----------
def _mkey(s):
    return re.sub(r"[^a-zа-яё0-9]", "", str(s or "").lower())


def model_map() -> dict:
    """Ключ-нормализация -> каноничное написание (самое частое)."""
    cnt = collections.defaultdict(collections.Counter)
    for c in inv.cams():
        if c.get("model"):
            cnt[_mkey(c["model"])][str(c["model"]).strip()] += 1
    try:
        d = read_all().get("Изменённые") or {}
        for row in d.get("rows", []):
            m = _cell(row, d["hdr"], "Модель")
            if m:
                cnt[_mkey(m)][str(m).strip()] += 1
    except Exception:
        log_exc("dq: model_map Изменённые")
    return {k: v.most_common(1)[0][0] for k, v in cnt.items()}


def model_variants() -> dict:
    """254: {канон: [варианты]} только там, где написаний больше одного."""
    cnt = collections.defaultdict(set)
    for c in inv.cams():
        if c.get("model"):
            cnt[_mkey(c["model"])].add(str(c["model"]).strip())
    try:
        d = read_all().get("Изменённые") or {}
        for row in d.get("rows", []):
            m = _cell(row, d["hdr"], "Модель")
            if m:
                cnt[_mkey(m)].add(str(m).strip())
    except Exception:
        pass
    mm = model_map()
    return {mm[k]: sorted(v) for k, v in cnt.items() if len(v) > 1}


# ---------- вспомогательные проверки ----------
def _ip_ok(v, policy):
    s = str(v or "").strip()
    if not IP_RE.match(s) or any(int(o) > 255 for o in s.split(".")):
        return False
    if policy == "any":
        return True
    subs = st.cget("cam_subnets") if policy == "cam" else st.cget("sw_subnets")
    return any(s.startswith(p + ".") for p in subs)


def _known_locs(data):
    """255: словарь допустимых локаций = частые значения + known из схемы."""
    cnt = collections.Counter()
    d = data.get("Все камеры") or {"hdr": [], "rows": []}
    for row in d["rows"]:
        for col in ("Расположение", "Объект"):
            v = _cell(row, d["hdr"], col)
            if v not in (None, ""):
                cnt[str(v).strip()] += 1
    known = {k for k, n in cnt.items() if n >= int(st.cget("loc_min_count"))}
    known |= set(ensure_schema().get("known_locations") or [])
    return known, cnt


def dup_serials() -> dict:
    """257: серийник -> отсортированный список разных IP (только где >1)."""
    ser = collections.defaultdict(set)
    try:
        d = read_all().get("Изменённые") or {}
        for row in d.get("rows", []):
            s, ip = _cell(row, d["hdr"], "Серийник"), _cell(row, d["hdr"], "Новый IP")
            if s and ip:
                ser[str(s).strip().upper()].add(str(ip).strip())
    except Exception:
        pass
    for ip, e in inv.fw_cache().items():
        if e.get("serial"):
            ser[str(e["serial"]).strip().upper()].add(ip)
    return {k: sorted(v) for k, v in ser.items() if len(v) > 1}


# ---------- 252 + 255-258, 260, 292, 293, 295: lint ----------
def lint(data=None) -> list:
    """Список проблем: {sheet,row,col,sev('crit'|'warn'),code,msg}."""
    schema = ensure_schema()
    data = data or read_all()
    issues = []

    def add(sheet, row, col, sev, code, msg):
        issues.append({"sheet": sheet, "row": row, "col": col,
                       "sev": sev, "code": code, "msg": msg})

    known_locs, _cnt = _known_locs(data)
    vlan_rules = dict(st.cget("vlan_rules") or {})
    oui = {p.upper() for p in st.cget("oui_whitelist") or []}
    for sheet, spec in schema["sheets"].items():
        d = data.get(sheet)
        if d is None:
            add(sheet, 0, "", "crit", "252", "лист отсутствует в файле")
            continue
        hdr = d["hdr"]
        for c in spec["columns"]:
            if c not in hdr:
                add(sheet, 1, c, "warn", "252", "колонки нет в листе")
        pref = tuple(p for p in spec.get("extra_prefixes", []) if p)
        for h in hdr:
            if h and h not in spec["columns"] and not h.startswith(pref):
                add(sheet, 1, h, "warn", "252", "колонка вне схемы")
            m = HDR_DATE_RE.search(h or "")
            if m and not DATE_RE.match(m.group(1)):
                add(sheet, 1, h, "warn", "295", "кривая дата в заголовке")
        for rn, row in enumerate(d["rows"], start=2):
            req_empty = all(_cell(row, hdr, k) in (None, "")
                            for k in spec.get("required", []))
            if req_empty:
                if any(c not in (None, "") for c in row):
                    add(sheet, rn, "", "warn", "292",
                        "мусорная строка: ключевые поля пусты, но данные есть")
                continue
            for c, meta in spec["columns"].items():
                v = _cell(row, hdr, c)
                if v in (None, ""):
                    continue
                t = meta.get("type")
                sv = str(v).strip()
                if t == "ip" and not _ip_ok(sv, meta.get("policy", "any")):
                    add(sheet, rn, c, "crit", "256", f"IP вне политики: {sv}")
                elif t == "mac":
                    canon = mac_canon(sv)
                    if not canon:
                        add(sheet, rn, c, "crit", "253", f"битый MAC: {sv}")
                    elif sv != canon:
                        add(sheet, rn, c, "warn", "253",
                            f"MAC не канон: {sv} -> {canon}")
                    elif oui and sheet in ("Все камеры", "Изменённые"):
                        model = str(_cell(row, hdr, "Модель камеры")
                                    or _cell(row, hdr, "Модель") or "").lower()
                        if (("apix" in model or "evidence" in model)
                                and canon[:8] not in oui):
                            add(sheet, rn, c, "warn", "258",
                                f"OUI {canon[:8]} не из белого списка вендора")
                elif t == "port" and not PORT_RE.match(sv):
                    add(sheet, rn, c, "warn", "252", f"порт не GE/TE/XGE: {sv}")
                elif t == "int" and not re.fullmatch(r"\d+", sv):
                    add(sheet, rn, c, "warn", "252", f"не целое: {sv}")
                elif t == "float":
                    try:
                        float(sv.replace(",", "."))
                    except ValueError:
                        add(sheet, rn, c, "warn", "252", f"не число: {sv}")
                elif t == "dt" and not DATE_RE.match(sv):
                    add(sheet, rn, c, "warn", "295",
                        f"дата не YYYY-MM-DD [HH:MM]: {sv}")
                elif t == "gw" and sv not in ("—", "-", ""):
                    if not (_ip_ok(sv, "cam") and sv.endswith(".254")):
                        add(sheet, rn, c, "warn", "256",
                            f"шлюз не .254 своей подсети: {sv}")
                elif t == "loc" and known_locs and sv not in known_locs:
                    add(sheet, rn, c, "warn", "255",
                        f"неизвестная локация: {sv}")
            if sheet == "Все камеры" and vlan_rules:
                obj = str(_cell(row, hdr, "Объект") or "")
                vl = str(_cell(row, hdr, "VLAN") or "").strip()
                for sub, allowed in vlan_rules.items():
                    if sub.lower() in obj.lower() and vl \
                            and vl not in [str(a) for a in allowed]:
                        add(sheet, rn, "VLAN", "warn", "293",
                            f"VLAN {vl} не по правилу «{sub}» {allowed}")
    # 260: ссылочная целостность камера<->коммутатор
    main, sw = data.get("Все камеры"), data.get("Лист1")
    if main and sw:
        sw_ips = {str(_cell(r, sw["hdr"], "ip VLAN 1") or "").strip()
                  for r in sw["rows"]} - {""}
        used = collections.Counter()
        for rn, row in enumerate(main["rows"], start=2):
            v = str(_cell(row, main["hdr"], "IP коммутатора") or "").strip()
            if v:
                used[v] += 1
                if v not in sw_ips:
                    add("Все камеры", rn, "IP коммутатора", "warn", "260",
                        f"коммутатора {v} нет в «Лист1»")
        for ip in sorted(sw_ips - set(used)):
            add("Лист1", 0, "ip VLAN 1", "warn", "260",
                f"коммутатор {ip} без единой камеры в инвентаре")
    # 257: дубли серийников
    for s, ips in dup_serials().items():
        add("Изменённые", 0, "Серийник", "crit", "257",
            f"серийник {s} у разных IP: {', '.join(ips)}")
    return issues


# ---------- 287: полнота данных + тренд ----------
def completeness(data=None) -> dict:
    """{'Все камеры': {'колонка': %, ...}, ...} — % заполненности."""
    data = data or read_all()
    out = {}
    for sheet, d in data.items():
        rows = d["rows"]
        if not rows:
            continue
        per = {}
        for i, h in enumerate(d["hdr"]):
            if not h:
                continue
            n = sum(1 for r in rows if i < len(r) and r[i] not in (None, ""))
            per[h] = round(100.0 * n / len(rows), 1)
        out[sheet] = per
    return out


def dq_history(limit=60) -> list:
    return store.jload(st.cget("dq_history_path"), [])[-limit:]


def note_dq(avg_score: float, compl: dict) -> None:
    """287: снимок метрик в _dq_history.json (не чаще раза в час)."""
    def _upd(hist):
        now = time.time()
        if hist and now - hist[-1].get("ts", 0) < 3600:
            return None
        hist.append({"ts": int(now),
                     "date": datetime.date.today().isoformat(),
                     "score": round(avg_score, 1),
                     "compl": {s: round(sum(p.values()) / max(len(p), 1), 1)
                               for s, p in compl.items()}})
        return hist[-365:]
    store.jupdate(st.cget("dq_history_path"), [], _upd)


# ---------- 288: DQ-score ----------
_W = (("ip", 15), ("mac", 15), ("model", 10), ("location", 10), ("switch", 10),
      ("port", 10), ("vlan", 5), ("cable", 5), ("status", 10), ("name", 5),
      ("facts", 5))


def dq_score_row(rec, facts_ok=None) -> int:
    """0-100 на строку «Все камеры»."""
    s = 0
    for k, w in _W:
        if k == "ip":
            ok = rec.get("ip") and _ip_ok(rec["ip"], "cam")
        elif k == "mac":
            ok = bool(mac_canon(rec.get("mac")))
        elif k == "facts":
            ok = facts_ok
            if ok is None and rec.get("mac"):
                ok = _facts_port_ok(rec)
        elif k == "status":
            ok = rec.get("status") not in (None, "")
        else:
            ok = rec.get(k) not in (None, "")
        if ok:
            s += w
    return s


def _facts_port_ok(rec) -> bool:
    try:
        for h in inv.switch_ports(rec.get("mac") or ""):
            if (str(h.get("sw_ip")) == str(rec.get("sw_ip"))
                    and str(h.get("port")) == str(rec.get("port"))):
                return True
    except Exception:
        pass
    return False


def dq_scores() -> list:
    """[(score, rec)] по всем камерам, худшие первыми."""
    res = [(dq_score_row(c), c) for c in inv.cams()]
    res.sort(key=lambda x: (x[0], str(x[1].get("ip"))))
    return res


# ---------- 300: сводка ----------
def summary() -> dict:
    """Агрегат для /dq: score, критичные/предупреждения, полнота, тренд."""
    data = read_all()
    issues = lint(data)
    scores = dq_scores()
    avg = round(sum(s for s, _ in scores) / max(len(scores), 1), 1)
    compl = completeness(data)
    prev = dq_history()
    note_dq(avg, compl)
    crit = [i for i in issues if i["sev"] == "crit"]
    by_code = collections.Counter(i["code"] for i in issues)
    return {"score": avg, "issues": issues, "crit": len(crit),
            "warn": len(issues) - len(crit), "by_code": dict(by_code),
            "compl": compl, "scores": scores,
            "prev": prev[-1] if prev else None,
            "variants": model_variants(), "mac_plan_n": len(mac_plan(data))}


# ---------- запись в xlsx с автобэкапом (253/273/298/259) ----------
def backup_xlsx(tag="dq") -> str:
    path = inv.inv_path()
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    bak = os.path.join(os.path.dirname(path), f"Все_камеры.backup.{ts}_{tag}.xlsx")
    shutil.copy2(path, bak)
    return bak


def apply_writes(writes, tag="dq", who="bot") -> str:
    """writes=[(лист, excel_row, колонка, старое, новое)] -> путь бэкапа.
    ПЕРЕД записью — обязательный бэкап; после — журнал 265 + сброс кэша."""
    if not writes:
        raise ValueError("нет изменений для записи")
    import openpyxl
    path = inv.inv_path()
    with inv.INV_WRITE_LOCK:  # C1: вся секция load→mutate→save
        bak = backup_xlsx(tag)
        wb = openpyxl.load_workbook(path)
        try:
            for sheet, rn, col, _old, new in writes:
                ws = wb[sheet]
                headers = [c.value for c in ws[1]]
                if col in headers:
                    ci = headers.index(col) + 1
                else:
                    ci = len(headers) + 1
                    ws.cell(row=1, column=ci, value=col)
                ws.cell(row=rn, column=ci, value=new)
            inv.save_wb(wb, path)
        finally:
            wb.close()
        with inv._lock:
            inv._inv["mtime"] = None
            inv._unk["mtime"] = None
    try:
        import bot_reconcile
        for sheet, rn, col, old, new in writes:
            bot_reconcile.record_change(who, sheet, f"row{rn}", col, old, new)
        bot_reconcile.after_xlsx_write(f"{tag}: {len(writes)} ячеек")
    except Exception:
        log_exc("dq: журнал изменений не записался")
    log(f"dq: записано {len(writes)} ячеек ({tag}), бэкап {os.path.basename(bak)}")
    return bak
