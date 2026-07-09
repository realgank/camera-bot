# -*- coding: utf-8 -*-
"""Волна F — команды качества данных: 252 /lint, 300 /dq (сводный отчёт +
xlsx с деталями), 253 /macfix (dry-run + кнопка «Применить» с бэкапом),
259+291 /migrate_plan (dry-run рефакторинга дат-колонок и UID-проверка;
сама миграция ТОЛЬКО по двухшаговой кнопке с автобэкапом), 294 /names
(пустые «Название (по ТЗ)» + генератор заготовок, БЕЗ автозаписи).
Прод-xlsx пишется только через bot_dq.apply_writes (бэкап обязателен)."""
import os
import re
import time
import tempfile
import threading
import collections

import bot_state as st
import bot_inventory as inv
import bot_dq as dq
import bot_store as store
from bot_tg import send, send_chunks, send_document, edit_message, answer_cq, chat_action
from bot_util import log, log_exc, esc, human_err

_PEND = {}      # token -> {"ts", "kind", "writes"/"plan"} — подтверждения
_pend_lock = threading.Lock()
_TTL = 300


def _stash(kind, payload) -> str:
    tok = format(int(time.time() * 1000) % 36 ** 6, "x")
    with _pend_lock:
        for k in [k for k, v in _PEND.items() if time.time() - v["ts"] > _TTL]:
            _PEND.pop(k, None)
        _PEND[tok] = {"ts": time.time(), "kind": kind, "payload": payload}
    return tok


def _take(tok, kind):
    with _pend_lock:
        e = _PEND.get(tok)
        if not e or e["kind"] != kind or time.time() - e["ts"] > _TTL:
            return None
        return _PEND.pop(tok)["payload"]


def _xlsx_bytes(build_fn) -> bytes:
    """openpyxl-книга через временный файл -> bytes."""
    import openpyxl
    wb = openpyxl.Workbook()
    build_fn(wb)
    fd, tmp = tempfile.mkstemp(suffix=".xlsx", prefix="dq_")
    os.close(fd)
    try:
        wb.save(tmp)
        with open(tmp, "rb") as f:
            return f.read()
    finally:
        try:
            os.unlink(tmp)
        except OSError:
            pass


# ---------- 252: /lint ----------
def cmd_lint(chat, arg="", reply_to=None):
    chat_action(chat)
    issues = dq.lint()
    if not issues:
        send(chat, "🧹 /lint: инвентарь чист по схеме "
                   "<code>inventory_schema.json</code> ✅", reply_to=reply_to)
        return
    crit = [i for i in issues if i["sev"] == "crit"]
    by = collections.Counter((i["code"], i["sev"]) for i in issues)
    lines = [f"🧹 <b>Lint инвентаря</b>: 🔴 {len(crit)} критических · "
             f"🟡 {len(issues) - len(crit)} предупреждений",
             "По классам: " + " · ".join(
                 f"[{c}]{'🔴' if s == 'crit' else ''}×{n}"
                 for (c, s), n in sorted(by.items()))]
    for i in (crit + [x for x in issues if x["sev"] != "crit"])[:25]:
        mark = "🔴" if i["sev"] == "crit" else "🟡"
        where = f"{i['sheet']}!{i['row']}" + (f" [{i['col']}]" if i["col"] else "")
        lines.append(f"{mark} {esc(where)}: {esc(i['msg'])}")
    if len(issues) > 25:
        lines.append(f"… и ещё {len(issues) - 25} — полный список в /dq (файлом)")
    lines.append("Нормализация MAC: /macfix · сводный балл: /dq")
    send_chunks(chat, lines)


# ---------- 300: /dq ----------
def _dq_details_xlsx(s) -> bytes:
    def build(wb):
        ws = wb.active
        ws.title = "Проблемы"
        ws.append(["Лист", "Строка", "Колонка", "Класс", "Серьёзность", "Что не так"])
        for i in s["issues"]:
            ws.append([i["sheet"], i["row"], i["col"], i["code"],
                       i["sev"], i["msg"]])
        ws2 = wb.create_sheet("DQ-score")
        ws2.append(["Score", "IP", "Название", "MAC", "Расположение", "Модель"])
        for sc, rec in s["scores"]:
            ws2.append([sc, rec.get("ip"), rec.get("name"), rec.get("mac"),
                        rec.get("location"), rec.get("model")])
        ws3 = wb.create_sheet("Полнота")
        ws3.append(["Лист", "Колонка", "% заполнено"])
        for sheet, per in s["compl"].items():
            for col, pct in per.items():
                ws3.append([sheet, col, pct])
        ws4 = wb.create_sheet("Модели-разнобой")
        ws4.append(["Канон", "Варианты"])
        for canon, vs in s["variants"].items():
            ws4.append([canon, " | ".join(vs)])
    return _xlsx_bytes(build)


def cmd_dq(chat, arg="", reply_to=None):
    chat_action(chat, "upload_document")
    send(chat, "🧪 Считаю качество данных (lint + score + полнота)…",
         silent=True, reply_to=reply_to)
    s = dq.summary()
    worst = s["scores"][:10]
    trend = ""
    if s["prev"]:
        d = s["score"] - s["prev"].get("score", s["score"])
        trend = (f" ({'📈+' if d > 0 else '📉'}{d:.1f} к {s['prev']['date']})"
                 if abs(d) >= 0.1 else f" (= {s['prev']['date']})")
    compl_main = s["compl"].get(inv.SHEET_MAIN) or {}
    avg_compl = round(sum(compl_main.values()) / max(len(compl_main), 1), 1)
    lines = [f"🧪 <b>Data Quality инвентаря</b>",
             f"⭐ Score парка: <b>{s['score']}/100</b>{trend}",
             f"🔴 критических: <b>{s['crit']}</b> · 🟡 предупреждений: "
             f"{s['warn']}",
             f"📊 полнота «Все камеры»: {avg_compl}% · "
             f"MAC к нормализации: {s['mac_plan_n']} (/macfix) · "
             f"моделей-разнобоев: {len(s['variants'])}"]
    if s["by_code"]:
        lines.append("Классы: " + " · ".join(
            f"[{c}]×{n}" for c, n in sorted(s["by_code"].items())))
    if worst:
        lines.append("\n🚨 <b>Топ-10 худших строк</b>:")
        for sc, rec in worst:
            lines.append(f"• {sc}/100 <code>{esc(rec.get('ip') or '—')}</code> "
                         f"{esc(rec.get('name') or '')} "
                         f"{esc(rec.get('location') or '')}")
    lines.append("\nДетали — документом ниже. /lint /macfix /migrate_plan /names")
    send_chunks(chat, lines)
    try:
        data = _dq_details_xlsx(s)
        send_document(chat, data, f"dq_{time.strftime('%Y%m%d_%H%M')}.xlsx",
                      caption=f"🧪 DQ-детали: {len(s['issues'])} проблем, "
                              f"score {s['score']}/100")
    except Exception as e:
        log_exc("/dq xlsx")
        send(chat, human_err("Файл деталей не собрался", e))


# ---------- 253: /macfix ----------
def cmd_macfix(chat, arg="", reply_to=None):
    chat_action(chat)
    plan = dq.mac_plan()
    if not plan:
        send(chat, "🧲 Все MAC во всех листах уже в каноне "
                   "<code>AA:BB:CC:DD:EE:FF</code> ✅", reply_to=reply_to)
        return
    lines = [f"🧲 <b>Нормализация MAC (dry-run)</b>: {len(plan)} ячеек"]
    for sheet, rn, col, old, new in plan[:20]:
        lines.append(f"• {esc(sheet)}!{rn}: <code>{esc(old)}</code> → "
                     f"<code>{esc(new)}</code>")
    if len(plan) > 20:
        lines.append(f"… и ещё {len(plan) - 20}")
    tok = _stash("macfix", plan)
    lines.append("\nЗапись — только по кнопке, с автобэкапом.")
    send_chunks(chat, lines)
    send(chat, f"Применить нормализацию {len(plan)} MAC?", silent=True,
         markup={"inline_keyboard": [[
             {"text": f"✅ Применить ({len(plan)})",
              "callback_data": f"dqmac:{tok}"},
             {"text": "✖️ Отмена", "callback_data": "cancel"}]]})


def cb_dqmac(chat, cq, tok):
    plan = _take(tok, "macfix")
    if plan is None:
        answer_cq(cq.get("id"), "⌛ План устарел — повтори /macfix")
        return
    answer_cq(cq.get("id"), "💾 Пишу с бэкапом…")
    try:
        bak = dq.apply_writes(plan, tag="macfix", who=chat)
    except Exception as e:
        log_exc("macfix apply")
        send(chat, human_err("Нормализация MAC не записалась", e))
        return
    send(chat, f"🧲 Нормализовано {len(plan)} MAC ✅\n"
               f"💾 Бэкап: <code>{esc(os.path.basename(bak))}</code>\n"
               f"Журнал: /history, срез CSV в exports\\ обновлён.")


# ---------- 259 + 291: /migrate_plan ----------
def _status_cols(data):
    hdr = (data.get(inv.SHEET_MAIN) or {}).get("hdr", [])
    return [h for h in hdr if h.startswith("Статус")]


def migrate_plan_info() -> dict:
    data = dq.read_all()
    scols = _status_cols(data)
    d = data.get(inv.SHEET_MAIN) or {"hdr": [], "rows": []}
    # 291: UID-проверка — уникальность и наличие MAC
    macs = collections.Counter()
    no_mac = 0
    for row in d["rows"]:
        m = dq.mac_canon(dq._cell(row, d["hdr"], "MAC-адрес"))
        if m:
            macs[m] += 1
        elif any(c not in (None, "") for c in row):
            no_mac += 1
    dups = {k: v for k, v in macs.items() if v > 1}
    n_events = 0
    for h in scols[:-1] if len(scols) > 1 else []:
        i = d["hdr"].index(h)
        n_events += sum(1 for r in d["rows"]
                        if i < len(r) and r[i] not in (None, ""))
    return {"scols": scols, "dups": dups, "no_mac": no_mac,
            "rows": len(d["rows"]), "n_events": n_events}


def cmd_migrate_plan(chat, arg="", reply_to=None):
    chat_action(chat)
    p = migrate_plan_info()
    scols = p["scols"]
    lines = ["🧭 <b>План миграции инвентаря (dry-run, БЕЗ записи)</b>",
             f"\n<b>259 — дат-колонки статусов</b>: найдено {len(scols)}:"]
    lines += [f"• <code>{esc(h)}</code>" for h in scols]
    if len(scols) > 1:
        last = scols[-1]
        lines.append(f"План: значения «{esc(last)}» → колонка «Статус», "
                     f"дата из заголовка → колонка «Проверено»; "
                     f"{p['n_events']} исторических значений из "
                     f"{len(scols) - 1} старых колонок → "
                     f"<code>_status_history.json</code> (296); "
                     f"старые колонки удаляются.")
    else:
        lines.append("Мигрировать нечего (одна колонка статуса или ноль).")
    lines.append(f"\n<b>291 — сквозной UID (MAC)</b>: строк {p['rows']}, "
                 f"без валидного MAC: <b>{p['no_mac']}</b>, "
                 f"дублей MAC: <b>{len(p['dups'])}</b>")
    for m, n in list(p["dups"].items())[:8]:
        lines.append(f"• дубль <code>{esc(m)}</code> ×{n}")
    if p["no_mac"] or p["dups"]:
        lines.append("⚠️ Сначала почини MAC (/macfix, /lint) — UID-связывание "
                     "листов будет неполным.")
    send_chunks(chat, lines)
    if len(scols) > 1:
        tok = _stash("migrate", scols)
        send(chat, "⚠️ Миграция меняет структуру прод-xlsx. Выполнить "
                   "с автобэкапом?", silent=True,
             markup={"inline_keyboard": [[
                 {"text": "🧭 Мигрировать…", "callback_data": f"dqmig:{tok}"},
                 {"text": "✖️ Отмена", "callback_data": "cancel"}]]})


def cb_dqmig(chat, cq, tok):
    """Шаг 2 — финальное подтверждение."""
    with _pend_lock:
        e = _PEND.get(tok)
    if not e or e["kind"] != "migrate" or time.time() - e["ts"] > _TTL:
        answer_cq(cq.get("id"), "⌛ План устарел — повтори /migrate_plan")
        return
    answer_cq(cq.get("id"), "Нужно финальное подтверждение")
    mid = (cq.get("message") or {}).get("message_id")
    txt = ("⚠️ <b>Финальное подтверждение миграции</b>\n"
           "Дат-колонки статусов схлопнутся в «Статус»+«Проверено», история "
           "уйдёт в _status_history.json. Бэкап делается автоматически.")
    kb = {"inline_keyboard": [[
        {"text": "✅ Да, мигрировать", "callback_data": f"dqmigy:{tok}"},
        {"text": "✖️ Отмена", "callback_data": "cancel"}]]}
    if mid:
        edit_message(chat, mid, txt, markup=kb)
    else:
        send(chat, txt, markup=kb)


def do_migrate(path=None) -> dict:
    """259: схлопывание дат-колонок (на path; прод — только из cb_dqmigy)."""
    import openpyxl
    path = path or inv.inv_path()
    wb = openpyxl.load_workbook(path)
    try:
        ws = wb[inv.SHEET_MAIN]
        hdr = [c.value for c in ws[1]]
        scols = [(i + 1, h) for i, h in enumerate(hdr)
                 if h and str(h).startswith("Статус")]
        if len(scols) < 2:
            return {"migrated": False, "reason": "меньше двух колонок статуса"}
        last_ci, last_h = scols[-1]
        m = dq.HDR_DATE_RE.search(str(last_h))
        checked = m.group(1) if m else ""
        events = []
        mac_ci = hdr.index("MAC-адрес") + 1 if "MAC-адрес" in hdr else None
        ip_ci = hdr.index("IP-адрес") + 1 if "IP-адрес" in hdr else None
        for ci, h in scols[:-1]:
            hm = dq.HDR_DATE_RE.search(str(h))
            hdate = hm.group(1) if hm else str(h)
            for rn in range(2, ws.max_row + 1):
                v = ws.cell(row=rn, column=ci).value
                if v in (None, ""):
                    continue
                events.append({
                    "ts": int(time.mktime(time.strptime(hdate, "%Y-%m-%d")))
                    if hm else 0,
                    "ip": str(ws.cell(row=rn, column=ip_ci).value or "")
                    if ip_ci else "",
                    "uid": inv.norm_mac(ws.cell(row=rn, column=mac_ci).value)
                    if mac_ci else "",
                    "st": str(v), "src": str(h)})
        ws.cell(row=1, column=last_ci, value="Статус")
        pc = len(hdr) + 1 if "Проверено" not in hdr \
            else hdr.index("Проверено") + 1
        ws.cell(row=1, column=pc, value="Проверено")
        for rn in range(2, ws.max_row + 1):
            if ws.cell(row=rn, column=last_ci).value not in (None, ""):
                ws.cell(row=rn, column=pc, value=checked)
        for ci, _h in sorted(scols[:-1], reverse=True):
            ws.delete_cols(ci)
        wb.save(path)
    finally:
        wb.close()
    if path == inv.inv_path():

        def _upd(data):
            data.setdefault("events", []).extend(events)
            return data
        store.jupdate(st.cget("status_history_path"),
                      {"cur": {}, "events": []}, _upd)
        with inv._lock:
            inv._inv["mtime"] = None
    return {"migrated": True, "dropped": len(scols) - 1,
            "events": len(events), "checked": checked}


def cb_dqmigy(chat, cq, tok):
    scols = _take(tok, "migrate")
    if scols is None:
        answer_cq(cq.get("id"), "⌛ Подтверждение устарело")
        return
    answer_cq(cq.get("id"), "🧭 Мигрирую с бэкапом…")
    try:
        bak = dq.backup_xlsx("migrate")
        r = do_migrate()
    except Exception as e:
        log_exc("migrate apply")
        send(chat, human_err("Миграция не прошла (файл цел, бэкап есть)", e))
        return
    if not r.get("migrated"):
        send(chat, f"🧭 Миграция не потребовалась: {esc(r.get('reason'))}")
        return
    try:
        import bot_reconcile
        bot_reconcile.record_change(chat, inv.SHEET_MAIN, "структура",
                                    "Статус-колонки",
                                    f"{r['dropped'] + 1} колонок", "Статус+Проверено")
        bot_reconcile.after_xlsx_write("миграция статус-колонок (259)")
    except Exception:
        pass
    log(f"migrate: -{r['dropped']} колонок, {r['events']} событий в историю")
    send(chat, f"🧭 <b>Миграция выполнена</b> ✅\n"
               f"Удалено старых колонок: {r['dropped']} · в историю статусов "
               f"ушло {r['events']} значений · «Проверено» = {esc(r['checked'])}\n"
               f"💾 Бэкап: <code>{esc(os.path.basename(bak))}</code>")


# ---------- 294: /names ----------
def cmd_names(chat, arg="", reply_to=None):
    import bot_zones
    cams = inv.cams()
    named = [c for c in cams if c.get("name")]
    empty = [c for c in cams if not c.get("name") and c.get("ip")]
    lines = [f"🏷 <b>«Название (по ТЗ)»</b>: заполнено {len(named)} из "
             f"{len(cams)} ({100 * len(named) // max(len(cams), 1)}%)"]
    if not empty:
        lines.append("Все камеры с IP поименованы ✅")
        send_chunks(chat, lines)
        return
    # заготовки: <SYS>-<корпус>.<NN> из локации/объекта + последний октет
    by_bld = collections.defaultdict(int)
    for c in named:
        p = bot_zones.parse_cam_name(c.get("name"))
        if p and p.get("bld"):
            by_bld[p["bld"]] += 1
    lines.append(f"Пустых: <b>{len(empty)}</b>. Заготовки (по шаблону "
                 f"SYS-<корпус>.<номер>, НЕ записываются):")
    for c in empty[:25]:
        fl = bot_zones.floor_of(c)
        base = re.sub(r"\D", "", str(c.get("ip") or "").rsplit(".", 1)[-1])
        bld = f"{fl}X" if fl is not None else "??"
        sugg = f"AS-{bld}.{int(base):02d}" if base else f"AS-{bld}.NN"
        lines.append(f"• <code>{c['ip']}</code> {esc(c.get('location') or '?')} "
                     f"→ <code>{esc(sugg)}</code>")
    if len(empty) > 25:
        lines.append(f"… и ещё {len(empty) - 25}")
    lines.append("Утверждённое имя пишется вручную: /note <ip> … или в xlsx.")
    send_chunks(chat, lines)


HANDLERS = {
    "/lint": cmd_lint, "/dq": cmd_dq, "/macfix": cmd_macfix,
    "/migrate_plan": cmd_migrate_plan, "/names": cmd_names,
}
ALIASES = {"/линт": "/lint", "/качество": "/dq", "/имена": "/names"}
CALLBACKS = {"dqmac": cb_dqmac, "dqmig": cb_dqmig, "dqmigy": cb_dqmigy}
