# -*- coding: utf-8 -*-
"""Волна J — I37+I38+I41: /provision — мастер ПНР заводской камеры Apix
(с завода все на 192.168.0.250 / Admin / 1234). Шаги: обнаружить → показать
модель/серийник/MAC (ONVIF) → /provision <целевой IP> (валидация I38: свободен
в инвентаре И не отвечает в сети) → двухшаговое подтверждение → смена IP
raw-SOAP SetNetworkInterfaces (камера уходит в ребут и поднимается на новом
IP) → шлюз ОТДЕЛЬНЫМ шагом на НОВОМ IP с ретраями (ловушка: HTTP 500 в окне
ребута — из памяти apix-factory-commissioning) → проверка доступности и
hwaddr → автозапись строки в xlsx с бэкапом (I41) + dirty для /sync.
Каждый шаг — в аудит. Пароли камер НИКОГДА не меняются."""
import re
import json
import time
import datetime

import bot_state as st
import bot_store as store
import bot_net as net
import bot_inventory as inv
import bot_onvifq as oq
import bot_sw_api as sw
from onvif_snap import DEV_NS, device_info, _grab
from bot_tg import send, edit_message, answer_cq, chat_action
from bot_util import log, log_exc, esc, human_err

SCH = "http://www.onvif.org/ver10/schema"
_confirm = sw.Confirm()

import threading
_await_name = {}          # chat_id -> (target_ip, ts): ждём название камеры
_name_lock = threading.Lock()


def _set_await_name(chat, target):
    with _name_lock:
        _await_name[chat] = (target, time.time())


def _pop_await_name(chat):
    ttl = float(st.cget("pending_ttl_s") or 600)
    with _name_lock:
        v = _await_name.pop(chat, None)
    if not v:
        return None
    target, ts = v
    return target if (time.time() - ts) <= ttl else None


def _audit(step, detail):
    log(f"provision: {step}: {detail}")
    try:
        import bot_obs
        bot_obs.audit(f"provision:{step}", detail, "")
    except Exception:
        pass


# ---------- лог провижининга: _provision_log.jsonl (успехи и отказы) ----------
def log_provision(rec):
    """Append-строка в _provision_log.jsonl (ts/dt добавляются автоматически)."""
    rec = {"ts": int(time.time()),
           "dt": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
           **rec}
    path = st.cget("provision_log_jsonl")
    try:
        with store.lock_for(path):
            with open(path, "a", encoding="utf-8") as f:
                f.write(json.dumps(rec, ensure_ascii=False) + "\n")
    except Exception:
        log_exc("provision: _provision_log.jsonl не записался")
    return rec


def provision_entries(limit=None):
    """Записи ПНР-лога (старые→новые). limit — только последние N."""
    path = st.cget("provision_log_jsonl")
    out = []
    try:
        with open(path, encoding="utf-8") as f:
            for line in f:
                try:
                    out.append(json.loads(line))
                except ValueError:
                    continue
    except OSError:
        return []
    return out[-limit:] if limit else out


def _logp(status, p, **extra):
    """Собрать и записать запись ПНР-лога по плану p и исходу status."""
    rec = {"status": status, "factory_ip": _fip(), "target": p.get("target"),
           "name": p.get("name") or "", "location": p.get("location") or "",
           "mac": p.get("hwaddr"), "model": p.get("model"),
           "serial": p.get("serial"), "gw": p.get("gw"),
           "prefix": p.get("prefix")}
    rec.update(extra)
    return log_provision(rec)


def _fip():
    return st.cget("health_factory_ip")


# ---------- парсер GetNetworkInterfaces (тестируется) ----------
def parse_ifaces(xml):
    """XML GetNetworkInterfaces -> {'token','hwaddr','addr','prefix','dhcp'}."""
    t = str(xml or "")
    m = re.search(r'<(?:\w+:)?NetworkInterfaces[^>]*\btoken="([^"]+)"', t)
    out = {"token": m.group(1) if m else None,
           "hwaddr": (_grab(t, "HwAddress") or "").upper(),
           "dhcp": (_grab(t, "DHCP") or "").strip().lower() == "true"}
    mm = re.search(r"<(?:\w+:)?Manual>(.*?)</(?:\w+:)?Manual>", t, re.S)
    src = mm.group(1) if mm else t
    out["addr"] = (_grab(src, "Address") or "").strip()
    try:
        out["prefix"] = int(_grab(src, "PrefixLength"))
    except (TypeError, ValueError):
        out["prefix"] = None
    return out


def get_ifaces(ip, timeout=8):
    """Живой GetNetworkInterfaces -> parse_ifaces (или {'error': …})."""
    r = oq.call(ip, "GetNetworkInterfaces",
                f'<GetNetworkInterfaces xmlns="{DEV_NS}"/>', timeout=timeout)
    if "text" not in r:
        return {"error": r.get("error") or "?"}
    return parse_ifaces(r["text"])


# ---------- I38: валидация целевого IP (чистая, тестируется) ----------
def target_check(ip_s, taken_ips, alive):
    """(ok, причина). taken_ips — IP из инвентаря; alive — отвечает ли адрес
    в сети (TCP/ping). Свободен = нет в инвентаре И молчит в сети."""
    if not net.valid_ip(ip_s):
        return False, "это не IPv4-адрес"
    last = int(ip_s.split(".")[-1])
    if last in (0, 255):
        return False, "сетевой/широковещательный адрес"
    if last == 254:
        return False, ".254 — шлюз подсети, занято по плану"
    if ip_s == _fip():
        return False, "это заводской адрес, целевым быть не может"
    if ip_s in taken_ips:
        return False, "занят в инвентаре"
    if alive:
        return False, "отвечает в сети (кем-то занят)"
    return True, ""


def _target_alive(ip):
    """I38: точно ли адрес молчит — TCP-проба + ping."""
    return net.tcp_alive(ip, ports=(80, 554, 8080), t=1.0) \
        or net.ping(ip) is not None


# ---------- raw-SOAP запись (только в рамках /provision) ----------
def _set_ip_body(token, new_ip, prefix):
    return (f'<SetNetworkInterfaces xmlns="{DEV_NS}">'
            f'<InterfaceToken>{token}</InterfaceToken>'
            f'<NetworkInterface>'
            f'<Enabled xmlns="{SCH}">true</Enabled>'
            f'<IPv4 xmlns="{SCH}"><Enabled>true</Enabled>'
            f'<Manual><Address>{new_ip}</Address>'
            f'<PrefixLength>{prefix}</PrefixLength></Manual>'
            f'<DHCP>false</DHCP></IPv4>'
            f'</NetworkInterface></SetNetworkInterfaces>')


def set_ip(old_ip, token, new_ip, prefix):
    """SetNetworkInterfaces: HTTP 200 = камера приняла и уходит в ребут."""
    r = oq.call(old_ip, "SetNetworkInterfaces",
                _set_ip_body(token, new_ip, prefix), timeout=10)
    if "text" in r and "fault" not in (r["text"] or "").lower():
        return True, ""
    return False, r.get("error") or _grab(r.get("text") or "", "Text") \
        or "SOAP Fault"


def set_gateway(ip, gw, retries=None, delay=None, sleep=time.sleep):
    """Шлюз ОТДЕЛЬНО, на новом IP, с ретраями: в окне ребута камера отдаёт
    HTTP 500 / рвёт соединение (память apix-factory-commissioning).
    Успех = верифицирован через GetNetworkDefaultGateway."""
    retries = retries or int(st.cget("provision_gw_retries"))
    delay = float(delay if delay is not None
                  else st.cget("provision_gw_delay_s"))
    body = (f'<SetNetworkDefaultGateway xmlns="{DEV_NS}">'
            f'<IPv4Address>{gw}</IPv4Address></SetNetworkDefaultGateway>')
    last = "?"
    for i in range(1, retries + 1):
        r = oq.call(ip, "SetNetworkDefaultGateway", body, timeout=8)
        last = r.get("error") or ""
        try:
            gnow = oq.get_net(ip)
            if (gnow.get("gateway") or "").strip() == gw:
                return True, i
            last = last or f"шлюз сейчас «{gnow.get('gateway')}»"
        except Exception as e:
            last = last or type(e).__name__
        if i < retries:
            sleep(delay)
    return False, f"{retries} попыток ({last})"


# ---------- I41: новая строка инвентаря с бэкапом ----------
def append_inventory_row(ip, mac, model, serial, note, name="", location=""):
    """Добавляет строку в «Все камеры» ПОСЛЕ автобэкапа. -> (бэкап, строка).
    name/location (если заданы) пишутся в «Название (по ТЗ)» и «Расположение»,
    чтобы новая камера была осмысленной и находилась по имени."""
    import openpyxl
    import bot_dq
    path = inv.inv_path()
    with inv.INV_WRITE_LOCK:  # C1: вся секция load→mutate→save
        bak = bot_dq.backup_xlsx("provision")
        n_max = 0
        for c in inv.cams():
            try:
                n_max = max(n_max, int(c.get("n") or 0))
            except (TypeError, ValueError):
                pass
        wb = openpyxl.load_workbook(path)
        try:
            ws = wb[inv.SHEET_MAIN]
            headers = [c.value for c in ws[1]]
            hi = {str(h): i + 1 for i, h in enumerate(headers) if h}
            rn = ws.max_row + 1
            vals = {"№": n_max + 1, "IP-адрес": ip, "MAC-адрес": mac,
                    "Модель камеры": model, "Примечание": note}
            if name:
                vals[inv.COLS["name"]] = name
            if location:
                vals[inv.COLS["location"]] = location
            if serial:
                for h in hi:
                    if "серийн" in h.lower():
                        vals[h] = serial
                        break
            for col, v in vals.items():
                if col in hi and v not in (None, ""):
                    ws.cell(row=rn, column=hi[col], value=v)
            inv.save_wb(wb, path)
        finally:
            wb.close()
        with inv._lock:
            inv._inv["mtime"] = None
    try:
        import bot_reconcile
        bot_reconcile.record_change("provision", inv.SHEET_MAIN, mac or ip,
                                    "IP-адрес", "", ip)
        bot_reconcile.after_xlsx_write(f"/provision {ip}")
    except Exception:
        log_exc("provision: журнал изменений не записался")
    return bak, rn


def _sync_now(reason):
    """Немедленный дифф-синк xlsx → Google-таблица сразу после провижина.
    -> (ok, строка-шаг). Провал — помечаем dirty (подхватит /autosync или
    ручной /sync)."""
    try:
        import bot_gsheets2
        bot_gsheets2.diff_sync(dry=False)
        try:
            import bot_autosync
            bot_autosync.clear_dirty()
        except Exception:
            pass
        return True, "✅ Синхронизировано с Google-таблицей"
    except Exception as e:
        log_exc("provision: немедленный синк упал")
        try:
            import bot_autosync
            bot_autosync.mark_dirty(reason)
        except Exception:
            pass
        return False, (f"⚠️ Синк в Google-таблицу не прошёл "
                       f"({esc(type(e).__name__)}) — помечено dirty, "
                       f"сделай /sync вручную")


def capture_name(chat, text):
    """Хук из bot_field.try_text: если ждём название камеры для ПНР — принять
    его и запустить план. -> True, если обработали."""
    target = _pop_await_name(chat)
    if not target:
        return False
    t = (text or "").strip()
    if t.lower() in ("-", "—", "без имени", "пропустить", "skip"):
        name, location = "", ""
    else:
        name, location = _parse_name_loc(t)
    _plan(chat, target, name, location)
    return True


# ---------- команда ----------
def _discover(chat, reply_to=None):
    fip = _fip()
    chat_action(chat)
    if not net.tcp_alive(fip, ports=(80,), t=1.5):
        send(chat, f"🏭 Заводская камера <code>{fip}</code> сейчас НЕ отвечает."
                   f"\nПодключи камеру (health-цикл сам заметит появление) и "
                   f"повтори /provision.", reply_to=reply_to)
        return
    info = device_info(fip, user=st.CAM_USER, pwd=st.CAM_PASS)
    ifc = get_ifaces(fip)
    if info.get("error") or ifc.get("error"):
        send(chat, human_err(f"<code>{fip}</code> отвечает по TCP, но ONVIF "
                             f"не дал данных",
                             info.get("error") or ifc.get("error")),
             reply_to=reply_to)
        return
    subs = st.cget("cam_subnets") or []
    send(chat,
         f"🏭 <b>Заводская камера найдена</b> <code>{fip}</code>\n"
         f"🎥 {esc(info.get('manufacturer'))} {esc(info.get('model'))} · "
         f"fw {esc(info.get('firmware'))}\n"
         f"#️⃣ sn {esc(info.get('serial'))} · MAC <code>{esc(ifc['hwaddr'])}</code>\n"
         f"⚠️ Если ARP и ONVIF-MAC различаются — на .250 сидит НЕСКОЛЬКО "
         f"камер (IP-конфликт): сначала разведи их по PoE.\n\n"
         f"Шаг 2 — целевой адрес: <code>/provision 10.20.50.123</code>\n"
         f"Свободные: " + " · ".join(f"/free_ip {s}" for s in subs[:4]),
         reply_to=reply_to)


def _parse_name_loc(rest):
    """'Название; Расположение' -> (name, location). Разделитель — ; или |."""
    rest = (rest or "").strip()
    if not rest:
        return "", ""
    for sep in (";", "|"):
        if sep in rest:
            a, b = rest.split(sep, 1)
            return a.strip(), b.strip()
    return rest, ""


def cmd_provision(chat, arg="", reply_to=None):
    a = (arg or "").strip()
    if not a:
        _discover(chat, reply_to)
        return
    parts = a.split(maxsplit=1)
    target = parts[0]
    name, location = _parse_name_loc(parts[1] if len(parts) > 1 else "")
    if not name:  # имя не задано в команде — спросим отдельным шагом
        _set_await_name(chat, target)
        send(chat,
             f"🏷 Как назвать камеру <code>{esc(target)}</code> в инвентаре?\n"
             f"Пришли <b>название (по ТЗ)</b>, можно с расположением через «;»:\n"
             f"<code>AS-7C.01; Лестница 5</code>\n"
             f"Или нажми «Без имени» — запишу только IP/MAC/модель.",
             markup={"inline_keyboard": [[
                 {"text": "⏭ Без имени", "callback_data": f"provskip:{target}"},
                 {"text": "✖️ Отмена", "callback_data": "cancel"}]]},
             reply_to=reply_to)
        return
    _plan(chat, target, name, location, reply_to)


def _plan(chat, target, name="", location="", reply_to=None):
    """Проверка целевого IP, опрос заводской камеры, построение плана ПНР
    и отправка первого подтверждения. name/location уйдут в строку xlsx."""
    fip = _fip()
    chat_action(chat)
    taken = {c["ip"] for c in inv.cams() if c.get("ip")}
    ok, why = target_check(target, taken, net.valid_ip(target)
                           and _target_alive(target))
    if not ok:
        pfx = target.rsplit(".", 1)[0] if net.valid_ip(target) else \
            (st.cget("cam_subnets") or ["10.20.50"])[0]
        send(chat, f"⛔ Целевой <code>{esc(target)}</code> не подходит: {why} "
                   f"(I38).\nСвободные адреса: /free_ip {pfx}",
             reply_to=reply_to)
        return
    if not net.tcp_alive(fip, ports=(80,), t=1.5):
        send(chat, f"🏭 Заводская <code>{fip}</code> не отвечает — нечего "
                   f"провижинить.", reply_to=reply_to)
        return
    info = device_info(fip, user=st.CAM_USER, pwd=st.CAM_PASS)
    ifc = get_ifaces(fip)
    if ifc.get("error") or not ifc.get("token"):
        send(chat, human_err("ONVIF заводской камеры не дал InterfaceToken",
                             ifc.get("error")), reply_to=reply_to)
        return
    gw = target.rsplit(".", 1)[0] + ".254"
    prefix = int(st.cget("provision_prefix_len"))
    key = f"{target}|{ifc['hwaddr']}"
    _confirm.put(key, {"target": target, "gw": gw, "prefix": prefix,
                       "token": ifc["token"], "hwaddr": ifc["hwaddr"],
                       "model": info.get("model") or "?",
                       "serial": info.get("serial") or "",
                       "name": name, "location": location})
    _audit("plan", f"{fip} -> {target}/{prefix} gw {gw} "
                   f"mac {ifc['hwaddr']} sn {info.get('serial')} "
                   f"name={name!r} loc={location!r}")
    nm_line = (f"🏷 <b>{esc(name)}</b>"
               + (f" · 📍 {esc(location)}" if location else "")
               + "\n") if name else ""
    send(chat,
         f"🛠 <b>ПНР: смена IP заводской камеры</b>\n"
         + nm_line +
         f"🎥 {esc(info.get('model'))} · sn {esc(info.get('serial'))} · "
         f"MAC <code>{esc(ifc['hwaddr'])}</code>\n"
         f"<code>{fip}</code> → <code>{target}</code>/{prefix}, "
         f"шлюз <code>{gw}</code>, DHCP off\n"
         f"После смены: проверка на новом IP, шлюз с ретраями (ловушка "
         f"HTTP 500), строка в xlsx с бэкапом (I41).\n"
         f"Подтверди в течение {st.cget('sw_confirm_ttl_s')}с:",
         markup={"inline_keyboard": [[
             {"text": "🛠 Сменить IP…", "callback_data": f"prov:{key}"},
             {"text": "✖️ Отмена", "callback_data": "cancel"}]]},
         reply_to=reply_to)


def cb_provskip(chat, cq, target):
    """Кнопка «Без имени»: планируем ПНР без названия/расположения."""
    _pop_await_name(chat)
    answer_cq(cq.get("id"), "Ок, без имени")
    _plan(chat, target, "", "")


def cb_prov(chat, cq, key):
    """Шаг 2: финальное подтверждение."""
    p = _confirm.take(key)
    if not p:
        answer_cq(cq.get("id"), "⌛ Подтверждение устарело — повтори /provision")
        return
    _confirm.put(key, p)
    answer_cq(cq.get("id"), "Нужно финальное подтверждение")
    mid = (cq.get("message") or {}).get("message_id")
    txt = (f"⚠️ <b>Финальное подтверждение ПНР</b>\n"
           f"<code>{_fip()}</code> → <code>{p['target']}</code> "
           f"(MAC {esc(p['hwaddr'])}) — камера перезагрузится. Жми ✅ в "
           f"течение {st.cget('sw_confirm_ttl_s')}с.")
    kb = {"inline_keyboard": [[
        {"text": "✅ Да, сменить IP", "callback_data": f"provy:{key}"},
        {"text": "✖️ Отмена", "callback_data": "cancel"}]]}
    if mid:
        edit_message(chat, mid, txt, markup=kb)
    else:
        send(chat, txt, markup=kb)


def cb_provy(chat, cq, key):
    p = _confirm.take(key)
    if not p:
        answer_cq(cq.get("id"), "⌛ Подтверждение устарело — повтори /provision")
        return
    answer_cq(cq.get("id"), "🛠 Меняю IP…")
    chat_action(chat)
    do_provision(chat, p)


def do_provision(chat, p):
    """Исполнение ПНР по подтверждённому плану p (шаги — в аудит)."""
    fip, target, gw = _fip(), p["target"], p["gw"]
    steps = []
    _audit("go", f"{fip} -> {target} подтверждён владельцем")
    ok, err = set_ip(fip, p["token"], target, p["prefix"])
    if not ok:
        _audit("set_ip_fail", f"{target}: {err}")
        _logp("ip_set_fail", p, detail=str(err)[:200])
        send(chat, human_err(f"SetNetworkInterfaces на <code>{fip}</code> "
                             f"не прошёл", err))
        return
    steps.append("✅ SetNetworkInterfaces принят — камера ушла в ребут")
    _audit("set_ip", f"{fip} -> {target} принят, жду на новом IP")
    # ждём камеру на новом IP
    wait_s = float(st.cget("provision_wait_s"))
    t0 = time.time()
    up = None
    while time.time() - t0 < wait_s:
        if net.tcp_alive(target, ports=(80,), t=1.0):
            up = time.time() - t0
            break
        time.sleep(3)
    if up is None:
        _audit("no_up", f"{target} не поднялась за {wait_s:.0f}s")
        _logp("no_up", p, wait_s=round(wait_s), detail="не поднялась на новом IP")
        send(chat, "\n".join(steps) + f"\n❌ Камера не поднялась на "
             f"<code>{target}</code> за {wait_s:.0f}с. Проверь /diag {target} "
             f"и /diag {fip}; строка в xlsx НЕ записана.")
        return
    steps.append(f"✅ Поднялась на <code>{target}</code> за {up:.0f}с")
    _audit("up", f"{target} за {up:.0f}s")
    # шлюз отдельным шагом с ретраями (ловушка HTTP 500 в окне ребута)
    gok, tries = set_gateway(target, gw)
    if gok:
        steps.append(f"✅ Шлюз <code>{gw}</code> установлен "
                     f"(с попытки {tries})")
        _audit("gateway", f"{target} gw {gw} с попытки {tries}")
    else:
        steps.append(f"⚠️ Шлюз НЕ подтвердился: {esc(str(tries))} — поставь "
                     f"вручную (камера при этом работает)")
        _audit("gateway_fail", f"{target}: {tries}")
    # верификация hwaddr — та ли камера поднялась
    ifc = get_ifaces(target)
    hw_ok = (ifc.get("hwaddr") or "").upper() == (p["hwaddr"] or "").upper()
    if not hw_ok:
        _audit("hwaddr_mismatch",
               f"{target}: ожидал {p['hwaddr']}, вижу {ifc.get('hwaddr')}")
        _logp("mac_mismatch", p, seen_mac=(ifc.get("hwaddr") or "").upper(),
              up_s=round(up, 1), gw_ok=gok, gw_tries=tries)
        steps.append(f"❌ MAC на новом IP <code>{esc(ifc.get('hwaddr') or '?')}"
                     f"</code> ≠ ожидаемому <code>{esc(p['hwaddr'])}</code> — "
                     f"на .250 было НЕСКОЛЬКО камер?! Строку в xlsx не пишу.")
        send(chat, "🛠 <b>ПНР: итог</b>\n" + "\n".join(steps))
        return
    steps.append(f"✅ MAC подтверждён: <code>{esc(p['hwaddr'])}</code>")
    # I41: строка в инвентарь (бэкап внутри) + немедленный синк
    rn = None
    synced = False
    row_ok = False
    try:
        note = (f"ПНР ботом {datetime.datetime.now():%d.%m.%Y %H:%M}"
                + ("" if gok else " (шлюз НЕ установлен!)"))
        bak, rn = append_inventory_row(target, p["hwaddr"], p["model"],
                                       p["serial"], note,
                                       name=p.get("name") or "",
                                       location=p.get("location") or "")
        row_ok = True
        steps.append(f"✅ Строка {rn} в xlsx (бэкап "
                     f"{esc(bak.rsplit(chr(92), 1)[-1])})")
        _audit("xlsx", f"{target} строка {rn}, бэкап {bak}")
        synced, sync_step = _sync_now(f"/provision {target}")
        steps.append(sync_step)
    except Exception as e:
        log_exc("provision: запись в xlsx")
        steps.append(f"❌ Строка в xlsx не записалась: {esc(type(e).__name__)}"
                     f" — добавь вручную")
    _logp("ok" if (row_ok and gok) else "ok_no_gw" if row_ok else "xlsx_fail",
          p, up_s=round(up, 1), gw_ok=gok, gw_tries=tries,
          xlsx_row=rn, synced=synced)
    send(chat, "🛠 <b>ПНР завершён</b>\n" + "\n".join(steps)
         + f"\nДальше: /accept {target} (приёмка), /sync (таблица).",
         markup={"inline_keyboard": [[
             {"text": "📸 Снимок", "callback_data": f"shot:{target}"},
             {"text": "🩺 Диаг", "callback_data": f"diag:{target}"}]]})


_LOG_ICON = {"ok": "✅", "ok_no_gw": "⚠️", "no_up": "❌", "mac_mismatch": "❌",
             "ip_set_fail": "❌", "xlsx_fail": "❌"}


def cmd_provlog(chat, arg="", reply_to=None):
    """Показать последние записи лога провижининга (_provision_log.jsonl)."""
    a = (arg or "").strip()
    n = max(1, min(int(a), 50)) if a.isdigit() else 10
    ents = provision_entries()
    if not ents:
        send(chat, "🧾 Лог ПНР пуст — ни одной камеры ещё не провижинили "
                   "(или файл _provision_log.jsonl не создан).",
             reply_to=reply_to)
        return
    lines = [f"🧾 <b>Лог провижининга</b> — последние "
             f"{min(n, len(ents))} из {len(ents)}:"]
    for e in ents[-n:][::-1]:
        ic = _LOG_ICON.get(e.get("status"), "·")
        nm = e.get("name") or "—"
        loc = f" · {esc(e['location'])}" if e.get("location") else ""
        gw = "" if e.get("gw_ok", True) else " · без шлюза"
        sy = "" if e.get("synced") else " · не синк"
        lines.append(f"{ic} <code>{esc(e.get('dt', '')[5:16])}</code> "
                     f"<code>{esc(e.get('target') or '?')}</code> "
                     f"{esc(nm)}{loc} · {esc(e.get('status'))}{gw}{sy}")
    lines.append("\nПодробно — в <code>_provision_log.jsonl</code>. "
                 "<code>/provlog N</code> — больше строк.")
    send(chat, "\n".join(lines), reply_to=reply_to)


HANDLERS = {"/provision": cmd_provision, "/provlog": cmd_provlog}
ALIASES = {"/пнр": "/provision", "/пнр_лог": "/provlog", "/провлог": "/provlog"}
CALLBACKS = {"prov": cb_prov, "provy": cb_provy, "provskip": cb_provskip}

try:  # /help: таб «⚙️ Операции» (Волна J)
    import bot_handlers_ux as _ux
    _ux.HELP_TABS["ops"] = ("⚙️ Операции", (
        "<b>Операции с парком</b> (Волна J, всё с двухшаговым "
        "подтверждением):\n"
        "/reboot <code>ip|имя</code> — PoE-ребут порта Cross-24 (отказ, если "
        "на порту >1 MAC)\n"
        "/reboot_soft — мягкий ONVIF-ребут\n"
        "/provision — ПНР заводской камеры 192.168.0.250: обнаружить,\n"
        "/provision <code>ip</code> — спросит название → сменить IP + шлюз + "
        "строка в xlsx (с именем) + сразу синк в Google-таблицу\n"
        "/provision <code>ip Название; Расположение</code> — то же, имя сразу\n"
        "/provlog <code>[N]</code> — лог провижининга (успехи и отказы)\n"
        "/macfill — заполнить пустые MAC в инвентаре (партиями по 10)\n"
        "/unknown_queue — очередь новых хостов → лист «Неизвестные»\n"
        "/autosync <code>on|off|N</code> — автосинк xlsx→Sheets раз в N часов\n"
        "/snapall <code>подсеть|зона</code> — снимки пачкой → Drive (дедуп)\n"
        "/clip <code>ip [сек]</code> — видеоклип с RTSP (нужен ffmpeg)\n"
        "Inline: <code>@бот запрос</code> в любом чате — карточки камер "
        "(включи у BotFather: /setinline; только для владельца)."))
except Exception:
    pass
