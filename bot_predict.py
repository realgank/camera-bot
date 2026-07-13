# -*- coding: utf-8 -*-
"""Волна G — предиктивная аналитика (317-321, 341-348):
317 rolling-baseline RTT (серия пингов ротацией, устойчивое ×3 = алерт);
318 джиттер (stdev серии); 319 контроль TTL (смена = другой маршрут,
ловушка proxy-ARP FortiGate); 320 /matrix — матрица частичной деградации
ICMP/TCP80/TCP554/ONVIF/snapshot/RTSP; 321 латентность ONVIF (метрика
из bot_camtime); 341 микро-ребуты (простой < 3 мин); 342 risk-score +
еженедельный топ-10; 343 /mtbf — MTBF/MTTR; 344 /sla — месячный xlsx
с учётом окон /maint; 345 /season — сезонность падений; 346 /heat —
теплокарта стабильности; 348 /trend — спарклайны ▁▂▃▅▇.
Данные — из SQLite bot_metrics (347). Read-only, пароли не меняются."""
import io
import re
import time
import datetime
import statistics
import subprocess
import collections
from concurrent.futures import ThreadPoolExecutor

import bot_state as st
import bot_net as net
import bot_inventory as inv
import bot_store as store
import bot_metrics as mx
from bot_tg import send, send_chunks, send_document, chat_action
from bot_util import log, log_exc, esc

BARS = "▁▂▃▄▅▆▇█"
MICRO_S = 180  # 341: простой короче = микро-ребут
BAD_EVENTS = ("frozen", "black", "snapdead", "rtsp_zombie", "clockjump",
              "sdp_drift", "port_new", "users_change", "arp_flap")
WD_NAMES = ["пн", "вт", "ср", "чт", "пт", "сб", "вс"]


# ---------- спарклайны (348) ----------
def spark(vals) -> str:
    vs = [v for v in vals if v is not None]
    if not vs:
        return ""
    lo, hi = min(vs), max(vs)
    rng = (hi - lo) or 1.0
    return "".join("·" if v is None else BARS[int((v - lo) / rng * 7)]
                   for v in vals)


def bucketize(series, n=24):
    """[(ts, val)] -> n медиан по равным интервалам времени (None=пусто)."""
    if not series:
        return []
    t0, t1 = series[0][0], series[-1][0]
    step = max(1.0, (t1 - t0) / n)
    buckets = [[] for _ in range(n)]
    for ts, v in series:
        buckets[min(n - 1, int((ts - t0) / step))].append(v)
    return [statistics.median(b) if b else None for b in buckets]


# ---------- 317-319: серия пингов ----------
def parse_ping(out: str):
    """(времена мс, TTL) из вывода ping Windows (RU/EN)."""
    times = [int(x) for x in re.findall(r"(?:time|время)[=<](\d+)",
                                        out or "", re.I)]
    ttls = re.findall(r"TTL=(\d+)", out or "", re.I)
    ttl = int(ttls[-1]) if ttls else None
    return times, ttl


def ping_series(ip: str, n: int = None):
    n = n or int(st.cget("pingq_count"))
    try:
        r = subprocess.run(
            ["ping", "-n", str(n), "-w", str(st.cget("ping_timeout_ms")), ip],
            timeout=st.cget("subproc_timeout_s") + n, **net._SUB_KW)
        return parse_ping(r.stdout)
    except Exception:
        return [], None


def _ping_one(ip: str) -> None:
    times, ttl = ping_series(ip)
    if not times:
        return
    med = statistics.median(times)
    jit = statistics.pstdev(times) if len(times) > 1 else 0.0
    base = mx.med(ip, "rtt", days=7, min_n=6)   # baseline ДО записи новой точки
    prev_ttl = mx.last_value(ip, "ttl")
    mx.metric_add(ip, "rtt", med)
    mx.metric_add(ip, "jit", jit)
    if ttl is not None:
        mx.metric_add(ip, "ttl", ttl)
        if prev_ttl is not None and int(prev_ttl) != ttl:   # 319
            if mx.event_add(ip, "ttl_change", f"{int(prev_ttl)} -> {ttl}"):
                mx.owner_alert(f"🛣 <b>TTL сменился</b>: "
                               f"{esc(inv.label(ip) or ip)} — было "
                               f"{int(prev_ttl)}, стало {ttl}. Ответ приходит "
                               f"другим маршрутом (proxy-ARP FortiGate?)",
                               aid="ttl_change")
    if base and base >= 1 and med >= base * 3:              # 317
        if mx.event_add(ip, "rtt_high", f"{med:.0f}ms vs base {base:.0f}ms"):
            mx.owner_alert(f"🐢 <b>RTT ×3 к баз. линии</b>: "
                           f"{esc(inv.label(ip) or ip)} — {med:.0f} мс при "
                           f"медиане 7 дн. {base:.0f} мс (линк/PoE "
                           f"деградирует?) · /trend {ip}", aid="rtt_spike")


def _tick_ping() -> None:
    if not st.cget("pingq_enabled"):
        return
    if not mx.due("pingq", float(st.cget("pingq_period_min")),
                  first_delay_s=360):
        return
    batch = mx.rotation_batch("pingq", int(st.cget("pingq_batch")))
    if not batch:
        return
    t0 = time.time()
    with ThreadPoolExecutor(max_workers=6, thread_name_prefix="pingq") as ex:
        list(ex.map(_ping_one, batch))
    log(f"pingq: серия пингов {len(batch)} камер за {time.time() - t0:.1f}s")


# ---------- 342: risk-score ----------
def risk_components(downs=0, micro=0, rtt_ratio=1.0, jitter=0.0,
                    events_n=0) -> dict:
    """Чистая функция: компоненты риска 0-100 (тестируется на синтетике)."""
    return {"флапы": min(40, int(downs) * 8),
            "микро-ребуты": min(20, int(micro) * 5),
            "RTT": 15 if rtt_ratio >= 3 else (8 if rtt_ratio >= 2 else 0),
            "джиттер": 5 if jitter >= 50 else 0,
            "события": min(20, int(events_n) * 5)}


def risk_total(comp: dict) -> int:
    return min(100, sum(comp.values()))


def risk_score(ip: str, downs_map=None, micro_map=None) -> tuple:
    """(балл, компоненты) по данным SQLite за 7 дней."""
    downs = (downs_map or mx.downs_count(7)).get(ip, 0)
    micro = (micro_map or mx.micro_reboots(7, MICRO_S)).get(ip, 0)
    base = mx.med(ip, "rtt", days=7, min_n=6)
    cur = mx.med(ip, "rtt", days=1, min_n=2)
    ratio = (cur / base) if base and cur and base >= 1 else 1.0
    jit = mx.last_value(ip, "jit") or 0.0
    ev = sum(1 for e in mx.events(ip=ip, days=7) if e["kind"] in BAD_EVENTS)
    comp = risk_components(downs, micro, ratio, jit, ev)
    return risk_total(comp), comp


def risk_top(n: int = 10) -> list:
    downs_map = mx.downs_count(7)
    micro_map = mx.micro_reboots(7, MICRO_S)
    ips = set(downs_map) | set(micro_map) | \
        {e["ip"] for e in mx.events(days=7) if e["kind"] in BAD_EVENTS}
    out = []
    for ip in ips:
        s, comp = risk_score(ip, downs_map, micro_map)
        if s > 0:
            out.append((s, ip, comp))
    return sorted(out, key=lambda x: -x[0])[:n]


def _risk_text(top) -> list:
    lines = ["🔮 <b>Прогноз отказов — топ риска</b> (7 дн., 342):"]
    if not top:
        lines.append("Кандидатов нет — парк спокоен ✅")
        return lines
    for s, ip, comp in top:
        why = ", ".join(f"{k}" for k, v in comp.items() if v)
        lines.append(f"• <b>{s}</b>/100 {esc(inv.label(ip) or ip)} — {esc(why)}"
                     f" · /trend {ip}")
    lines.append("Микро-ребуты (341) = простои &lt; 3 мин: watchdog камеры "
                 "перезапускает зависшую прошивку.")
    return lines


def cmd_risk(chat, arg="", reply_to=None):
    chat_action(chat)
    send_chunks(chat, _risk_text(risk_top(10)))


def _tick_risk() -> None:
    """342: еженедельный топ-10 владельцу."""
    now = datetime.datetime.now()
    if now.weekday() != int(st.cget("risk_weekday")) \
            or now.hour < int(st.cget("risk_hour")):
        return
    wk = now.strftime("%G-%V")
    if mx.kv_get("risk_week") == wk:
        return
    mx.kv_set("risk_week", wk)
    top = risk_top(10)
    if top:
        mx.owner_alert("\n".join(_risk_text(top)), silent=True, aid="risk_digest")


# ---------- 343: MTBF/MTTR ----------
def mtbf_mttr(ip: str, days: float = 30) -> dict:
    now = time.time()
    t0 = now - days * 86400
    downs = mx.downs_count(days, ip).get(ip, 0)
    dt = mx.downtime_s(ip, t0, now)
    return {"downs": downs, "downtime": dt,
            "mtbf_h": ((days * 86400 - dt) / downs / 3600) if downs else None,
            "mttr_min": (dt / downs / 60) if downs else None}


def cmd_mtbf(chat, arg="", reply_to=None):
    chat_action(chat)
    a = (arg or "").strip()
    ip = a if net.valid_ip(a) else (inv.resolve_ip(a) if a else None)
    if ip:
        m = mtbf_mttr(ip)
        send(chat, f"⚙️ <b>{esc(inv.label(ip) or ip)}</b> за 30 дн. (343):\n"
                   f"падений: <b>{m['downs']}</b> · простой: "
                   f"{m['downtime'] / 3600:.1f} ч\n"
                   + (f"MTBF: <b>{m['mtbf_h']:.0f} ч</b> · MTTR: "
                      f"<b>{m['mttr_min']:.0f} мин</b>" if m["downs"]
                      else "Отказов не было — MTBF/MTTR не определены ✅"),
             reply_to=reply_to)
        return
    rows = []
    for cip, n in mx.downs_count(30).items():
        if n:
            m = mtbf_mttr(cip)
            rows.append((m["mtbf_h"] or 0, cip, m))
    rows.sort(key=lambda x: x[0])
    lines = ["⚙️ <b>MTBF/MTTR за 30 дн.</b> — худшие по наработке (343):"]
    if not rows:
        lines.append("Падений за период нет ✅")
    for _k, cip, m in rows[:12]:
        lines.append(f"• {esc(inv.label(cip) or cip)}: MTBF "
                     f"{m['mtbf_h']:.0f} ч · MTTR {m['mttr_min']:.0f} мин "
                     f"· {m['downs']} пад.")
    send_chunks(chat, lines)


# ---------- 344: /sla ----------
def _maint_windows_for(ip: str, t0: float, t1: float) -> list:
    """Окна /maint (159), затрагивающие камеру, зажатые в [t0, t1]."""
    out = []
    try:
        d = store.jload(st.cget("maint_path"),
                        {"seq": 0, "active": [], "log": []})
        for w in (d.get("active") or []) + (d.get("log") or []):
            if ip not in (w.get("ips") or []):
                continue
            a = max(t0, w.get("start") or 0)
            b = min(t1, w.get("ended") or w.get("until") or 0)
            if b > a:
                out.append((a, b))
    except Exception:
        log_exc("sla: maint windows")
    return out


def sla_month(year: int, month: int) -> dict:
    """Доступность парка/подсетей за месяц с учётом окон работ (344)."""
    t0 = datetime.datetime(year, month, 1).timestamp()
    nxt = (datetime.datetime(year + 1, 1, 1) if month == 12
           else datetime.datetime(year, month + 1, 1))
    t1 = min(time.time(), nxt.timestamp())
    period = max(1.0, t1 - t0)
    try:
        import bot_health
        ips = sorted(bot_health.snapshot()["ips"])
    except Exception:
        ips = []
    per_ip, by_sub = {}, collections.defaultdict(lambda: [0, 0.0])
    for ip in ips:
        dt = mx.downtime_s(ip, t0, t1, exclude=_maint_windows_for(ip, t0, t1))
        per_ip[ip] = dt
        sub = ip.rsplit(".", 1)[0]
        by_sub[sub][0] += 1
        by_sub[sub][1] += dt
    total_dt = sum(per_ip.values())
    avail = 100.0 * (1 - total_dt / (period * max(1, len(ips))))
    return {"t0": t0, "t1": t1, "period": period, "ips": len(ips),
            "avail": avail, "per_ip": per_ip,
            "by_sub": {s: (n, 100.0 * (1 - d / (period * n)))
                       for s, (n, d) in by_sub.items()}}


def cmd_sla(chat, arg="", reply_to=None):
    chat_action(chat, "upload_document")
    m = re.match(r"^(\d{4})-(\d{1,2})$", (arg or "").strip())
    if m:
        year, month = int(m.group(1)), int(m.group(2))
    else:
        today = datetime.date.today()
        year, month = today.year, today.month
    r = sla_month(year, month)
    if not r["ips"]:
        send(chat, "SLA: нет данных health-надзора.", reply_to=reply_to)
        return
    import openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SLA"
    ws.append([f"SLA-отчёт за {year}-{month:02d} (окна /maint исключены)",
               time.strftime("%Y-%m-%d %H:%M")])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append(["Камер в надзоре", r["ips"]])
    ws.append(["Доступность парка, %", round(r["avail"], 3)])
    ws.append([])
    ws.append(["Подсеть", "Камер", "Доступность, %"])
    for s, (n, pct) in sorted(r["by_sub"].items()):
        ws.append([s + ".x", n, round(pct, 3)])
    ws2 = wb.create_sheet("Худшие камеры")
    ws2.append(["IP", "Название", "Простой, ч", "Доступность, %"])
    worst = sorted(r["per_ip"].items(), key=lambda kv: -kv[1])[:30]
    for ip, dt in worst:
        if dt <= 0:
            continue
        ws2.append([ip, str((inv.get(ip) or {}).get("name") or ""),
                    round(dt / 3600, 2),
                    round(100.0 * (1 - dt / r["period"]), 3)])
    buf = io.BytesIO()
    wb.save(buf)
    send_document(chat, buf.getvalue(), f"SLA_{year}-{month:02d}.xlsx",
                  caption=f"📊 SLA {year}-{month:02d}: парк "
                          f"<b>{r['avail']:.2f}%</b> · {r['ips']} камер · "
                          + " · ".join(f"{s}.x {pct:.1f}%"
                                       for s, (_n, pct)
                                       in sorted(r["by_sub"].items())))


# ---------- 345: /season ----------
def season_hist(days: float = 30):
    """(по часам 0-23, по дням недели 0-6) из down-переходов."""
    by_h = collections.Counter()
    by_wd = collections.Counter()
    for e in mx.transitions(days=days):
        if e["ev"] != "down":
            continue
        lt = time.localtime(e["ts"])
        by_h[lt.tm_hour] += 1
        by_wd[lt.tm_wday] += 1
    return by_h, by_wd


def cmd_season(chat, arg="", reply_to=None):
    try:
        days = max(1, min(int(arg), 90))
    except (TypeError, ValueError):
        days = 30
    by_h, by_wd = season_hist(days)
    total = sum(by_h.values())
    if not total:
        send(chat, f"📅 Падений за {days} дн. не зафиксировано ✅",
             reply_to=reply_to)
        return
    mx_h = max(by_h.values())
    lines = [f"📅 <b>Сезонность падений за {days} дн.</b> ({total} шт., 345):",
             "<pre>по часам:"]
    for h in range(24):
        n = by_h.get(h, 0)
        lines.append(f"{h:02d}  {'█' * round(n / mx_h * 20)} {n or ''}")
    mx_w = max(by_wd.values())
    lines.append("\nпо дням недели:")
    for i, nm in enumerate(WD_NAMES):
        n = by_wd.get(i, 0)
        lines.append(f"{nm}  {'█' * round(n / mx_w * 20)} {n or ''}")
    lines.append("</pre>Пики в одни часы = питание/уборка по расписанию, "
                 "а не случайность.")
    send(chat, "\n".join(lines), reply_to=reply_to)


# ---------- 346: /heat ----------
def cmd_heat(chat, arg="", reply_to=None):
    chat_action(chat)
    prefix = (arg or "").strip().rstrip(".")
    try:
        import bot_health
        snap = bot_health.snapshot()["ips"]
    except Exception:
        snap = {}
    if not snap:
        send(chat, "Нет данных health-надзора.", reply_to=reply_to)
        return
    if not prefix:
        lines = ["🌡 <b>Теплокарта стабильности за 7 дн.</b> (346):"]
        subs = sorted({ip.rsplit(".", 1)[0] for ip in snap})
        for sub in subs:
            ips = [ip for ip in snap if ip.startswith(sub + ".")]
            pcts = [mx.uptime_pct(ip, 7) for ip in ips]
            bad = sum(1 for p in pcts if p < 99)
            avg = sum(pcts) / len(pcts)
            mark = "🟩" if avg >= 99 else ("🟨" if avg >= 95 else "🟥")
            lines.append(f"{mark} <code>{sub}.x</code>: {avg:.1f}% · "
                         f"камер {len(ips)} · с проблемами {bad} · "
                         f"/heat {sub}")
        send(chat, "\n".join(lines), reply_to=reply_to)
        return
    if not net.valid_prefix(prefix):
        send(chat, "Формат: <code>/heat 10.20.50</code>", reply_to=reply_to)
        return
    cells = []
    worst = []
    for i in range(256):
        ip = f"{prefix}.{i}"
        if ip not in snap:
            cells.append("⬛")
            continue
        p = mx.uptime_pct(ip, 7)
        cells.append("🟩" if p >= 99 else ("🟨" if p >= 95 else "🟥"))
        if p < 99:
            worst.append((p, ip))
    grid = "\n".join("".join(cells[r * 16:(r + 1) * 16]) for r in range(16))
    txt = (f"🌡 <b>{esc(prefix)}.x — аптайм за 7 дн.</b> (строки по 16)\n{grid}\n"
           f"🟩 ≥99% · 🟨 ≥95% · 🟥 &lt;95% · ⬛ не в надзоре")
    for p, ip in sorted(worst)[:8]:
        txt += f"\n🔻 <code>{ip}</code> {p:.1f}% — {esc(inv.label(ip) or '')}"
    send(chat, txt, reply_to=reply_to)


# ---------- 348: /trend ----------
def cmd_trend(chat, arg="", reply_to=None):
    a = (arg or "").strip()
    ip = a if net.valid_ip(a) else (inv.resolve_ip(a) if a else None)
    if not ip:
        send(chat, "Спарклайны камеры: <code>/trend 10.20.50.51</code>",
             reply_to=reply_to)
        return
    chat_action(chat)
    lines = [f"📈 <b>Тренды {esc(inv.label(ip) or ip)}</b> (348):"]
    up = mx.daily_uptime(ip, 14)
    lines.append(f"аптайм 14 дн.: <code>{spark(up)}</code> "
                 f"{up[-1]:.1f}% сегодня")
    for kind, nm, unit, days in (("rtt", "RTT", "мс", 7),
                                 ("jit", "джиттер", "мс", 7),
                                 ("onvif_ms", "ONVIF", "мс", 7),
                                 ("snap_kb_day", "кадр (день)", "КБ", 30),
                                 ("clock_off", "часы", "с", 30),
                                 ("rtsp_kbps", "битрейт", "кбит/с", 30)):
        ser = mx.series(ip, kind, days)
        if len(ser) < 2:
            continue
        b = bucketize(ser, 24)
        vals = [v for v in b if v is not None]
        lines.append(f"{nm} {days}д: <code>{spark(b)}</code> "
                     f"{min(vals):.0f}…{max(vals):.0f} {unit}, "
                     f"сейчас {ser[-1][1]:.0f}")
    ev = mx.events(ip=ip, days=7)
    if ev:
        lines.append("события 7 дн.: " + ", ".join(
            e["kind"] for e in ev[-8:]))
    if len(lines) == 2 and not ev:
        lines.append("Метрик пока мало — ротации (пинги/часы/снимки) "
                     "накопят за сутки.")
    send_chunks(chat, lines)


# ---------- 320: /matrix ----------
def cmd_matrix(chat, arg="", reply_to=None):
    a = (arg or "").strip()
    ip = a if net.valid_ip(a) else (inv.resolve_ip(a) if a else None)
    if not ip:
        send(chat, "Матрица деградации: <code>/matrix 10.20.50.51</code>",
             reply_to=reply_to)
        return
    chat_action(chat)
    send(chat, f"🧬 Проверяю все слои <code>{ip}</code> …",
         reply_to=reply_to, silent=True)
    rows = []
    rows.append(("ICMP ping", net.ping(ip) is not None))
    rows.append(("TCP 80", net.tcp_alive(ip, ports=(80,), t=1.0)))
    rows.append(("TCP 554", net.tcp_alive(ip, ports=(554,), t=1.0)))
    from onvif_snap import device_info, get_snapshot
    info = device_info(ip, user=st.CAM_USER, pwd=st.CAM_PASS)
    rows.append(("ONVIF", bool(info.get("model"))))
    data, _m = get_snapshot(ip, user=st.CAM_USER, pwd=st.CAM_PASS)
    rows.append(("snapshot", bool(data)))
    import bot_rtsp
    r = bot_rtsp.describe(ip)
    rows.append(("RTSP SDP", bool(r.get("ok"))))
    table = "\n".join(f"{nm:<9} {'OK' if ok else '--'}" for nm, ok in rows)
    n_ok = sum(1 for _n, ok in rows if ok)
    verdict = ("✅ все слои живы" if n_ok == len(rows) else
               ("❌ камера полностью недоступна" if n_ok == 0 else
                f"⚠️ частичная деградация: {n_ok}/{len(rows)} слоёв (320) — "
                f"«полуживая» камера"))
    send(chat, f"🧬 <b>{esc(inv.label(ip) or ip)}</b>\n<pre>{esc(table)}</pre>"
               f"{verdict}")


def _tick() -> None:
    _tick_ping()
    _tick_risk()


try:
    import bot_health as _bh
    _bh.MINUTE_TICKS.append(_tick)
except Exception:
    pass

HANDLERS = {"/risk": cmd_risk, "/mtbf": cmd_mtbf, "/sla": cmd_sla,
            "/season": cmd_season, "/heat": cmd_heat, "/trend": cmd_trend,
            "/matrix": cmd_matrix}
ALIASES, CALLBACKS = {"/риск": "/risk", "/тренд": "/trend"}, {}
