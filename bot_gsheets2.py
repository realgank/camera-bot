# -*- coding: utf-8 -*-
"""Волна I — дифф-синк xlsx -> Google Sheets и журнал событий: 429 дифф
вместо clear+rewrite (batchGet -> сравнение -> точечный values.batchUpdate;
форматирование и заметки не затираются), 430 одна пакетная запись, 432 лок
(локальный + developerMetadata с TTL), 433 /sync dry с кнопкой «Применить»,
434 детект чужих правок (revisions.list), 435 обратный синк «Комментария»
(код+тесты, живьём НЕ выполняется), 436 контроль лимита ячеек, 413 автодифф
в лист «Изменённые (бот)», 405 версия синка в developerMetadata и на
«Дашборде», 402 лист «Журнал событий» (хук owner_alert, буфер + тик).
Прод-xlsx здесь НЕ пишется (apply_comments — вручную/тесты на копии)."""
import re
import json
import time
import datetime
import threading

import bot_state as st
import google_api as g
import bot_sheets as sh
from bot_util import log, log_exc, esc, BOT_VERSION

SHEETS = "https://sheets.googleapis.com/v4/spreadsheets"
META_KEY = "camera_bot_sync"        # 405
LOCK_KEY = "camera_bot_sync_lock"   # 432

_DRY = {"ts": 0.0}                  # 433: свежесть dry-run для кнопки
_ev_lock = threading.Lock()
_EV_BUF: list = []                  # 402: буфер строк журнала событий
_ev_sheet = [None]                  # кэш «лист журнала существует»


# ---------- чистые функции (тестируются на фикстурах) ----------
def col_letter(n):
    """0 -> A, 25 -> Z, 26 -> AA."""
    s = ""
    n += 1
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def norm_cell(v):
    """Нормализация ячейки для сравнения xlsx против UNFORMATTED_VALUE."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def _pad(row, w):
    out = [("" if c is None else c) for c in row[:w]]
    return out + [""] * (w - len(out))


def grid_diff(remote, local):
    """429: сравнение сеток. Возвращает {'blocks': [(row0, rows)], 'cells',
    'rows', 'extra_rows', 'width'}. Ширина — по xlsx: колонки таблицы правее
    (ручные, HYPERLINK 411) диффом НЕ трогаются."""
    w = max((len(r) for r in local), default=0)
    changed, cells = [], 0
    for i, lrow in enumerate(local):
        rrow = remote[i] if i < len(remote) else []
        d = 0
        for j in range(w):
            lv = norm_cell(lrow[j]) if j < len(lrow) else ""
            rv = norm_cell(rrow[j]) if j < len(rrow) else ""
            if lv != rv:
                d += 1
        if d:
            changed.append(i)
            cells += d
    blocks = []
    for i in changed:
        if blocks and i == blocks[-1][0] + len(blocks[-1][1]):
            blocks[-1][1].append(_pad(local[i], w))
        else:
            blocks.append([i, [_pad(local[i], w)]])
    return {"blocks": [(a, rows) for a, rows in blocks], "cells": cells,
            "rows": len(changed), "extra_rows": max(0, len(remote) - len(local)),
            "width": w}


def row_details(remote, local, key_col, limit=300):
    """413: [(ключ, поле, было, стало)] по изменившимся строкам (кроме шапки)."""
    hdr = local[0] if local else []
    w = max((len(r) for r in local), default=0)
    out = []
    for i in range(1, len(local)):
        rrow = remote[i] if i < len(remote) else []
        lrow = local[i]
        key = str(lrow[key_col]) if (key_col is not None and key_col < len(lrow)
                                     and lrow[key_col] != "") else f"строка {i + 1}"
        for j in range(w):
            lv = norm_cell(lrow[j]) if j < len(lrow) else ""
            rv = norm_cell(rrow[j]) if j < len(rrow) else ""
            if lv != rv:
                f = str(hdr[j]) if j < len(hdr) and hdr[j] else col_letter(j)
                out.append((key, f, rv[:80], lv[:80]))
                if len(out) >= limit:
                    return out
    return out


def pull_comments(remote_rows, col_name="Комментарий", ip_col_name="IP-адрес"):
    """435 (pure): пары (ip, комментарий) из живой сетки главного листа."""
    if not remote_rows:
        return []
    hdr = [str(h) for h in remote_rows[0]]
    try:
        ci = hdr.index(col_name)
        ii = hdr.index(ip_col_name)
    except ValueError:
        return []
    out = []
    for row in remote_rows[1:]:
        ip = str(row[ii]).strip() if ii < len(row) else ""
        cm = str(row[ci]).strip() if ci < len(row) else ""
        if ip and cm:
            out.append((ip, cm))
    return out


def apply_comments(pairs, xlsx_path, col_name="Комментарий"):
    """435: перенос комментариев в xlsx (с бэкапом). ЖИВЬЁМ НЕ ВЫЗЫВАЕТСЯ —
    только вручную/тестами на копии. Возвращает число записанных."""
    import shutil
    import openpyxl
    import bot_inventory as inv
    bak = xlsx_path.replace(".xlsx", "") + \
        f".backup.{time.strftime('%Y%m%d_%H%M%S')}_comm.xlsx"
    shutil.copy2(xlsx_path, bak)
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[inv.SHEET_MAIN]
    hdr = [c.value for c in ws[1]]
    try:
        ci = hdr.index(col_name) + 1
    except ValueError:
        ci = len(hdr) + 1
        ws.cell(row=1, column=ci, value=col_name)
    try:
        ii = [str(h) for h in hdr].index("IP-адрес") + 1
    except ValueError:
        wb.close()
        return 0
    by_ip = dict(pairs)
    n = 0
    for r in range(2, ws.max_row + 1):
        ip = str(ws.cell(row=r, column=ii).value or "").strip()
        if ip in by_ip and str(ws.cell(row=r, column=ci).value or "") != by_ip[ip]:
            ws.cell(row=r, column=ci, value=by_ip[ip])
            n += 1
    wb.save(xlsx_path)
    wb.close()
    return n


# ---------- метаданные, лок, ревизии ----------
def _meta(sid, sa):
    return g.gjson("GET", f"{SHEETS}/{sid}", sa_path=sa, timeout=30,
                   fields="sheets.properties(sheetId,title,gridProperties"
                          "(rowCount,columnCount)),developerMetadata"
                          "(metadataId,metadataKey,metadataValue)")


def _meta_get(meta, key):
    for dm in meta.get("developerMetadata") or []:
        if dm.get("metadataKey") == key:
            return dm
    return None


def _meta_del(sid, sa, key):
    try:
        g.gjson("POST", f"{SHEETS}/{sid}:batchUpdate", sa_path=sa, json={
            "requests": [{"deleteDeveloperMetadata": {"dataFilter": {
                "developerMetadataLookup": {"metadataKey": key}}}}]})
    except Exception:
        pass  # ключа не было — не страшно


def _meta_set(sid, sa, key, value):
    _meta_del(sid, sa, key)
    g.gjson("POST", f"{SHEETS}/{sid}:batchUpdate", sa_path=sa, json={
        "requests": [{"createDeveloperMetadata": {"developerMetadata": {
            "metadataKey": key, "metadataValue": str(value)[:1000],
            "visibility": "DOCUMENT", "location": {"spreadsheet": True}}}}]})


def last_sync_meta(meta):
    """405: {'ts':…, 'ver':…, 'rows':…} из developerMetadata или {}."""
    dm = _meta_get(meta, META_KEY)
    try:
        return json.loads(dm["metadataValue"]) if dm else {}
    except (ValueError, KeyError):
        return {}


def foreign_edit(sid, sa, since_ts):
    """434: (когда, кто) последней чужой ревизии после since_ts или None."""
    j = g.gjson("GET",
                f"https://www.googleapis.com/drive/v3/files/{sid}/revisions",
                sa_path=sa, scope=g.SCOPE_DRIVE, timeout=30,
                fields="revisions(modifiedTime,lastModifyingUser(emailAddress))",
                params={"pageSize": 1000})
    revs = j.get("revisions") or []
    if not revs:
        return None
    last = revs[-1]
    mt = last.get("modifiedTime") or ""
    try:
        ts = datetime.datetime.strptime(
            mt.split(".")[0].rstrip("Z"), "%Y-%m-%dT%H:%M:%S") \
            .replace(tzinfo=datetime.timezone.utc).timestamp()
    except ValueError:
        return None
    who = (last.get("lastModifyingUser") or {}).get("emailAddress") or "?"
    if ts > float(since_ts or 0) + 120 and who != g.sa_email(sa):
        return (mt, who)
    return None


def cells_total(meta):
    """436: суммарное число ячеек spreadsheet по gridProperties."""
    n = 0
    for s in meta.get("sheets") or []:
        gp = (s.get("properties") or {}).get("gridProperties") or {}
        n += int(gp.get("rowCount") or 0) * int(gp.get("columnCount") or 0)
    return n


# ---------- 429/430: дифф-синк ----------
def _batch_get(sid, sa, titles):
    params = [("valueRenderOption", "UNFORMATTED_VALUE"),
              ("dateTimeRenderOption", "FORMATTED_STRING")] + \
             [("ranges", f"'{t}'") for t in titles]
    j = g.gjson("GET", f"{SHEETS}/{sid}/values:batchGet", sa_path=sa,
                params=params, timeout=120)
    out = {}
    for t, vr in zip(titles, j.get("valueRanges") or []):
        out[t] = vr.get("values") or []
    return out


def diff_sync(dry=False, force=False, want_snapshot=False):
    """429: дифф-синк. Возвращает текст-отчёт. dry — только посчитать (433)."""
    if not sh.SYNC_LOCK.acquire(blocking=False):
        return ("⏳ Синхронизация уже идёт — подожди.", None)
    t0 = time.time()
    locked_meta = False
    sid, sa = st.cget("sheet_id"), st.cget("sa_path")
    try:
        sheets, offline = sh._load_xlsx()
        meta = _meta(sid, sa)
        # 432: чужой лок в developerMetadata
        lk = _meta_get(meta, LOCK_KEY)
        ttl = float(st.cget("sync_lock_ttl_min")) * 60
        if lk and not force:
            try:
                age = time.time() - float(lk.get("metadataValue") or 0)
            except ValueError:
                age = ttl + 1
            if age < ttl:
                return (f"🔒 Таблицу сейчас синкает другой процесс "
                        f"(лок {age:.0f}s). Повтори позже или /sync force.", None)
        # 436: лимит ячеек
        total_cells = cells_total(meta)
        warn_cells = ""
        if total_cells > int(st.cget("cells_limit_warn")):
            warn_cells = (f"\n⚠️ В таблице ~{total_cells:,} ячеек — близко к "
                          f"лимиту 10 млн. Заархивируй старые листы.")
        ev_rows = next(
            (int((s["properties"].get("gridProperties") or {}).get("rowCount") or 0)
             for s in meta.get("sheets") or []
             if s["properties"]["title"] == st.cget("sheet_events_name")), 0)
        if ev_rows > int(st.cget("sheet_events_max_rows")):
            warn_cells += (f"\n⚠️ «{st.cget('sheet_events_name')}» разросся до "
                           f"{ev_rows} строк — пора перенести в архив.")
        # 434: чужие ручные правки после нашего последнего синка
        prev = last_sync_meta(meta)
        fe = None
        try:
            fe = foreign_edit(sid, sa, prev.get("ts_epoch"))
        except Exception:
            log_exc("sync: revisions.list (не критично)")
        existing = {s["properties"]["title"]: s["properties"]["sheetId"]
                    for s in meta["sheets"]}
        add = [{"addSheet": {"properties": {"title": t}}}
               for t in sheets if t not in existing]
        if add and not dry:
            g.gjson("POST", f"{SHEETS}/{sid}:batchUpdate", sa_path=sa,
                    json={"requests": add})
            meta2 = _meta(sid, sa)
            existing = {s["properties"]["title"]: s["properties"]["sheetId"]
                        for s in meta2["sheets"]}
        remote = _batch_get(sid, sa, [t for t in sheets if t in existing])
        import bot_inventory as inv
        diffs, data, clears, details = {}, [], [], []
        for title, local in sheets.items():
            rem = remote.get(title, [])
            d = grid_diff(rem, local)
            diffs[title] = d
            for r0, rows in d["blocks"]:
                data.append({"range": f"'{title}'!A{r0 + 1}:"
                                      f"{col_letter(d['width'] - 1)}{r0 + len(rows)}",
                             "values": rows})
            if d["extra_rows"]:
                rw = max([len(r) for r in rem] + [d["width"]])
                clears.append(f"'{title}'!A{len(local) + 1}:"
                              f"{col_letter(rw - 1)}{len(rem)}")
            if title == inv.SHEET_MAIN and d["rows"]:
                hdr = [str(h) for h in (local[0] if local else [])]
                key_col = hdr.index("IP-адрес") if "IP-адрес" in hdr else None
                details = row_details(rem, local, key_col)
        n_cells = sum(d["cells"] for d in diffs.values())
        n_rows = sum(d["rows"] for d in diffs.values())
        n_extra = sum(d["extra_rows"] for d in diffs.values())
        stats = [f"«{t}»: ✏️{d['cells']} ячеек в {d['rows']} строках"
                 + (f", −{d['extra_rows']} лишних строк" if d["extra_rows"] else "")
                 for t, d in diffs.items() if d["cells"] or d["extra_rows"]]
        fe_s = (f"\n⚠️ ручные правки в таблице: {esc(fe[1])} · {esc(fe[0])}"
                if fe else "")
        if dry:  # 433
            _DRY["ts"] = time.time()
            kb = {"inline_keyboard": [[
                {"text": "✅ Применить", "callback_data": "gsapply:go"},
                {"text": "✖️ Отмена", "callback_data": "cancel"}]]} \
                if (n_cells or n_extra or add) else None
            return (("🧪 <b>Dry-run синка</b> (ничего не записано)\n"
                     + ("\n".join(stats) if stats else "✅ различий нет")
                     + (f"\n➕ новых листов: {len(add)}" if add else "")
                     + f"\nитого: {n_cells} ячеек · {n_rows} строк"
                     + fe_s + warn_cells), kb)
        # 426: снапшот перед большим синком / при чужих правках
        snap_s = "не требовался"
        if want_snapshot or fe or n_cells > 2000 or n_extra > 50:
            try:
                import bot_gdrive2
                snap = bot_gdrive2.snapshot_spreadsheet()
                snap_s = f"сделан ({snap.get('name')})"
            except Exception:
                log_exc("sync: снапшот-копия не удалась")
                snap_s = "НЕ УДАЛСЯ (пишем без него)"
        # 432: наш лок в developerMetadata
        try:
            _meta_set(sid, sa, LOCK_KEY, time.time())
            locked_meta = True
        except Exception:
            log_exc("sync: не смог поставить метадата-лок (не критично)")
        written = 0
        if data:  # 430: одна пакетная запись
            r = g.gjson("POST", f"{SHEETS}/{sid}/values:batchUpdate", sa_path=sa,
                        json={"valueInputOption": "RAW", "data": data}, timeout=180)
            written = int(r.get("totalUpdatedCells") or 0)
        if clears:
            g.gjson("POST", f"{SHEETS}/{sid}/values:batchClear", sa_path=sa,
                    json={"ranges": clears}, timeout=60)
        reqs = sh.fmt_requests(existing, sheets, offline)
        if reqs:
            g.gjson("POST", f"{SHEETS}/{sid}:batchUpdate", sa_path=sa,
                    json={"requests": reqs}, timeout=120)
        if details:  # 413
            try:
                now_row = time.strftime("%Y-%m-%d %H:%M:%S")
                _append_rows(st.cget("sync_diff_sheet"),
                             [[now_row, k, f, o, n] for k, f, o, n in details],
                             header=["Когда", "IP", "Поле", "Было", "Стало"])
            except Exception:
                log_exc("sync: автодифф в «Изменённые (бот)» не записался")
        try:  # Волна F (285): KPI-лист
            sh._push_kpi(sid, sa, sheets)
        except Exception:
            log_exc("sync: KPI-лист (не критично)")
        # 405: шапка версии на «Дашборде» + developerMetadata
        now_s = time.strftime("%Y-%m-%d %H:%M:%S")
        ver = int(prev.get("ver") or 0) + 1
        main_rows = len(sheets.get(inv.SHEET_MAIN) or [])
        try:
            from urllib.parse import quote
            g.gjson("PUT", f"{SHEETS}/{sid}/values/"
                    f"{quote(st.cget('kpi_sheet_name') + '!A8')}"
                    f"?valueInputOption=RAW", sa_path=sa, timeout=30,
                    json={"values": [[f"Синк: {now_s} · v{ver} · {main_rows} строк"
                                      f" · дифф {n_cells} ячеек"]]})
        except Exception:
            log_exc("sync: шапка версии (не критично)")
        try:
            _meta_set(sid, sa, META_KEY, json.dumps(
                {"ts": now_s, "ts_epoch": time.time(), "ver": ver,
                 "rows": main_rows, "cells": n_cells, "bot": BOT_VERSION},
                ensure_ascii=False))
        except Exception:
            log_exc("sync: developerMetadata (не критично)")
        fmt_s = ""
        if st.cget("sheet_fmt_enabled"):  # 401/403-411 best-effort
            try:
                import bot_gfmt
                done = bot_gfmt.decorate(sid, sa, sheets, offline)
                fmt_s = f"\n🎨 оформление: {esc(', '.join(done))}" if done else ""
            except Exception:
                log_exc("sync: оформление (не критично)")
        dt = time.time() - t0
        sh.save_sync_state({"ts": now_s, "cells": n_cells, "offline": len(offline),
                            "mode": "diff", "ver": ver, "written": written,
                            "sheets": {t: len(r) for t, r in sheets.items()},
                            "sec": round(dt, 1)})
        note_event("sync", "", "", f"дифф-синк v{ver}: {n_cells} ячеек, "
                                   f"{n_rows} строк, {dt:.1f}s")
        log(f"gsync: дифф {n_cells} ячеек / {n_rows} строк (записано {written}), "
            f"{dt:.1f}s")
        return (("✅ <b>Дифф-синк готов</b>\n"
                 + ("\n".join(stats) if stats else "✅ различий не было")
                 + f"\nитого: изменено {n_cells} ячеек в {n_rows} строках"
                 f" · записано {written}"
                 + f"\n🔴 офлайн подсвечено: {len(offline)}"
                 + f"\n📸 снапшот-копия: {esc(snap_s)}" + fe_s + warn_cells + fmt_s
                 + f"\n⏱ {dt:.1f}s · <a href=\"{sh.sheet_url()}\">таблица</a>"),
                None)
    finally:
        if locked_meta:
            try:
                _meta_del(sid, sa, LOCK_KEY)
            except Exception:
                pass
        sh.SYNC_LOCK.release()


def _append_rows(title, rows, header=None):
    """values.append в лист (создаёт лист с шапкой при отсутствии)."""
    from urllib.parse import quote
    sid, sa = st.cget("sheet_id"), st.cget("sa_path")
    meta = g.gjson("GET", f"{SHEETS}/{sid}", sa_path=sa, timeout=30,
                   fields="sheets.properties.title")
    titles = {s["properties"]["title"] for s in meta.get("sheets") or []}
    if title not in titles:
        g.gjson("POST", f"{SHEETS}/{sid}:batchUpdate", sa_path=sa,
                json={"requests": [{"addSheet": {"properties": {"title": title}}}]})
        if header:
            rows = [header] + rows
    g.gjson("POST", f"{SHEETS}/{sid}/values/{quote(title + '!A1')}:append"
                    f"?valueInputOption=RAW&insertDataOption=INSERT_ROWS",
            sa_path=sa, json={"values": rows}, timeout=60)
    return len(rows)


# ---------- 402: журнал событий ----------
def note_event(kind, ip="", name="", details=""):
    """Строка в буфер журнала (сбрасывается минутным тиком)."""
    if not st.cget("sheet_events_enabled"):
        return
    with _ev_lock:
        if len(_EV_BUF) < 200:
            _EV_BUF.append([time.strftime("%Y-%m-%d %H:%M:%S"), str(kind),
                            str(ip), str(name)[:80], str(details)[:300]])


def _alert_hook(text):
    plain = re.sub(r"<[^>]+>", "", str(text)).strip()
    note_event("alert", "", "", plain.split("\n", 1)[0][:200])


def _events_tick():
    with _ev_lock:
        if not _EV_BUF:
            return
        rows, _EV_BUF[:] = list(_EV_BUF), []
    try:
        _append_rows(st.cget("sheet_events_name"), rows,
                     header=["Когда", "Тип", "IP", "Имя", "Детали"])
    except Exception:
        log_exc("gsheets2: журнал событий не записался")
        with _ev_lock:  # вернём строки, не потеряем (с потолком буфера)
            _EV_BUF[:] = (rows + _EV_BUF)[:200]


# ---------- команды ----------
def cmd_sync(chat, arg="", reply_to=None):
    """429/433: /sync [dry|full|force|comments]."""
    from bot_tg import send, chat_action
    a = (arg or "").strip().lower()
    chat_action(chat)
    try:
        if a == "full" or not st.cget("sync_diff_enabled"):
            send(chat, sh.sync(), reply_to=reply_to)
            return
        if a == "comments":  # 435: только показать план (xlsx НЕ пишется)
            rem = _batch_get(st.cget("sheet_id"), st.cget("sa_path"),
                             [__import__("bot_inventory").SHEET_MAIN])
            pairs = pull_comments(list(rem.values())[0] if rem else [],
                                  col_name=st.cget("sheet_comment_col"))
            txt = (f"💬 Комментариев в таблице: {len(pairs)}\n"
                   + "\n".join(f"• <code>{esc(ip)}</code>: {esc(c[:60])}"
                               for ip, c in pairs[:20])
                   + ("\n…" if len(pairs) > 20 else "")
                   + "\nЗапись в xlsx не выполняется (защита 435).")
            send(chat, txt, reply_to=reply_to)
            return
        send(chat, "🔄 Считаю дифф xlsx ↔ таблица …", silent=True,
             reply_to=reply_to)
        txt, kb = diff_sync(dry=(a == "dry"), force=(a == "force"))
        send(chat, txt, markup=kb, reply_to=reply_to)
    except Exception as e:
        log_exc("/sync (волна I)")
        from bot_util import human_err
        send(chat, human_err("Синк упал", e), reply_to=reply_to)


def cb_apply(chat, cq, payload):
    """433: кнопка «Применить» после dry-run."""
    from bot_tg import send, answer_cq
    if time.time() - _DRY["ts"] > 600:
        answer_cq(cq.get("id"), "⌛ Dry-run устарел — повтори /sync dry")
        return
    answer_cq(cq.get("id"), "▶️ Применяю…")
    txt, _kb = diff_sync(dry=False)
    send(chat, txt)


HANDLERS = {"/sync": cmd_sync}
ALIASES = {"/синк": "/sync"}
CALLBACKS = {"gsapply": cb_apply}

try:  # 402: подписка на алерты + минутный тик сброса буфера
    import bot_metrics as _mx
    _mx.ALERT_HOOKS.append(_alert_hook)
    import bot_health as _bh
    _bh.MINUTE_TICKS.append(_events_tick)
except Exception:
    log_exc("gsheets2: не смог зарегистрировать хук/тик")

try:  # /help: таб «☁️ Google» (bot_handlers_ux на лимите 500 строк)
    import bot_handlers_ux as _ux
    _ux.HELP_TABS["gg"] = ("☁️ Google", (
        "<b>Google-экосистема</b> (Волна I):\n"
        "/sync — дифф-синк xlsx→таблица (заметки не затираются)\n"
        "/sync dry — посчитать, кнопка «Применить» · /sync full — перезапись\n"
        "/sync force — сквозь лок · /sync comments — колонка «Комментарий»\n"
        "/sheet — ссылка · /sheet_fmt — оформить · /accept_sheet — приёмка\n"
        "/snaps <code>ip</code> — снимки камеры в Drive · /snaprotate — ротация\n"
        "/report_pdf — «Дашборд» PDF · /gsnap — снапшот-копии таблицы\n"
        "/ga_health — сервис-аккаунт: доступы, квота, ключ, probe-задача\n"
        "/gcal — календарь ППР и напоминания (нужен Calendar API)\n"
        "/nightly — ночные задачи: сверка 03:00, бэкап xlsx в Drive,\n"
        "git-автокоммит, чистка · /sla_gs — месячная SLA-таблица\n"
        "Алерты бота дублируются в лист «Журнал событий»."))
except Exception:
    pass
