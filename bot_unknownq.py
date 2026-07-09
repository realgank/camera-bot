# -*- coding: utf-8 -*-
"""Волна J — I36: автопополнение листа «Неизвестные устройства».
Новые хосты из /find и ночной сверки, которых нет в инвентаре (ни по IP, ни
по MAC), копятся в _unknown_queue.json (note_hosts — чистая логика в
filter_new). Раз в сутки (unknown_daily_hour) бот присылает сводку с кнопкой
«Записать N в лист» — запись в xlsx ТОЛЬКО по кнопке, с двухшаговым
подтверждением и автобэкапом. /unknown_queue — посмотреть очередь вручную."""
import time
import datetime
import threading

import bot_state as st
import bot_net as net
import bot_inventory as inv
import bot_store as store
import bot_sw_api as sw
from bot_tg import send, edit_message, answer_cq
from bot_util import log, log_exc, esc, human_err

_confirm = sw.Confirm()
UNK_HDRS = ["MAC-адрес", "Вендор (OUI)", "Категория", "Коммутатор",
            "IP коммутатора", "Порт", "VLAN", "Примечание"]


def _qpath():
    return st.cget("unknown_queue_path")


def filter_new(hosts, inv_ips, inv_nmacs, sw_prefixes, factory_ip):
    """Чистая: hosts={ip: mac|'—'} -> {key: {ip, mac}} только настоящих
    «неизвестных» (нет в инвентаре по IP и MAC, не свитч-подсеть, не шлюз,
    не заводской IP). key = нормализованный MAC либо IP."""
    out = {}
    for ip, mac in (hosts or {}).items():
        if ip in inv_ips or ip == factory_ip:
            continue
        if ip.endswith(".254") or ip.endswith(".255"):
            continue
        if any(ip.startswith(p + ".") for p in sw_prefixes):
            continue
        nm = inv.norm_mac(mac) if mac and mac != "—" else ""
        if nm and nm in inv_nmacs:
            continue
        out[nm or ip] = {"ip": ip, "mac": (mac if mac and mac != "—" else "")}
    return out


def note_hosts(hosts, source="find"):
    """Хук из /find и ночной сверки: скопить новые хосты (best-effort)."""
    try:
        inv_ips = {c["ip"] for c in inv.cams() if c.get("ip")}
        inv_nmacs = {c["nmac"] for c in inv.cams() if c.get("nmac")}
        fresh = filter_new(hosts, inv_ips, inv_nmacs,
                           st.cget("sw_subnets") or [],
                           st.cget("health_factory_ip"))
        if not fresh:
            return 0
        now_s = time.strftime("%Y-%m-%d %H:%M")

        def upd(q):
            cut = time.time() - float(st.cget("unknown_keep_days")) * 86400
            for key, e in fresh.items():
                cur = q.get(key) or {"first": now_s, "seen": 0, "src": source}
                cur.update(ip=e["ip"], last=now_s, last_ts=time.time(),
                           seen=int(cur.get("seen") or 0) + 1)
                if e["mac"]:
                    cur["mac"] = e["mac"]
                q[key] = cur
            return {k: v for k, v in q.items()
                    if k.startswith("_") or (v.get("last_ts") or 0) >= cut}
        store.jupdate(_qpath(), {}, upd)
        log(f"unknownq: +{len(fresh)} хостов из {source}")
        return len(fresh)
    except Exception:
        log_exc("unknownq: note_hosts")
        return 0


def queue():
    """Актуальная очередь (без служебных _-ключей), заново фильтрованная
    против инвентаря (вдруг уже записали руками)."""
    q = store.jload(_qpath(), {})
    inv_ips = {c["ip"] for c in inv.cams() if c.get("ip")}
    inv_nmacs = {c["nmac"] for c in inv.cams() if c.get("nmac")}
    out = {}
    for k, e in q.items():
        if k.startswith("_") or not isinstance(e, dict):
            continue
        if e.get("ip") in inv_ips:
            continue
        if e.get("mac") and inv.norm_mac(e["mac"]) in inv_nmacs:
            continue
        out[k] = e
    return out


def _existing_unk_macs():
    hdr, rows = inv.unknown_devices()
    if "MAC-адрес" not in hdr:
        return set()
    i = hdr.index("MAC-адрес")
    return {inv.norm_mac(r[i]) for r in rows if i < len(r) and r[i]}


def summary_text(q):
    lines = [f"❔ <b>Очередь «Неизвестных»</b>: {len(q)} новых хостов "
             f"(нет в инвентаре):"]
    for k, e in sorted(q.items(), key=lambda kv: kv[1].get("ip") or "")[:15]:
        mac = e.get("mac") or "MAC не пойман"
        lines.append(f"• <code>{esc(e.get('ip') or '?')}</code> · "
                     f"<code>{esc(mac)}</code> {esc(net.vendor(e.get('mac')))}"
                     f" · виден {e.get('seen', 1)}× (с {esc(e.get('first') or '?')})")
    if len(q) > 15:
        lines.append(f"… и ещё {len(q) - 15}")
    return "\n".join(lines)


def _offer(chat, q):
    tok = str(int(time.time()))
    _confirm.put("unk" + tok, True)
    send(chat, summary_text(q) + "\nЗапись в xlsx — только по кнопке, "
                                 "с автобэкапом.",
         markup={"inline_keyboard": [[
             {"text": f"💾 Записать {len(q)} в «Неизвестные устройства»…",
              "callback_data": f"unkw:{tok}"},
             {"text": "✖️ Отмена", "callback_data": "cancel"}]]})


def cmd_unknown_queue(chat, arg="", reply_to=None):
    q = queue()
    if not q:
        send(chat, "❔ Очередь «Неизвестных» пуста — новые хосты копятся из "
                   "/find и ночной сверки.", reply_to=reply_to)
        return
    _offer(chat, q)


def cb_unkw(chat, cq, tok):
    if _confirm.take("unk" + tok) is None:
        answer_cq(cq.get("id"), "⌛ Сводка устарела — повтори /unknown_queue")
        return
    _confirm.put("unky" + tok, True)
    answer_cq(cq.get("id"), "Нужно финальное подтверждение")
    n = len(queue())
    mid = (cq.get("message") or {}).get("message_id")
    txt = (f"⚠️ <b>Финальное подтверждение</b>: дописать {n} строк в лист "
           f"«Неизвестные устройства» (перед записью — автобэкап)?")
    kb = {"inline_keyboard": [[
        {"text": "✅ Да, записать", "callback_data": f"unkwy:{tok}"},
        {"text": "✖️ Отмена", "callback_data": "cancel"}]]}
    if mid:
        edit_message(chat, mid, txt, markup=kb)
    else:
        send(chat, txt, markup=kb)


def cb_unkwy(chat, cq, tok):
    if _confirm.take("unky" + tok) is None:
        answer_cq(cq.get("id"), "⌛ Подтверждение устарело — повтори /unknown_queue")
        return
    answer_cq(cq.get("id"), "💾 Записываю…")
    try:
        n, bak = write_unknowns()
    except Exception as e:
        log_exc("unknownq: запись")
        send(chat, human_err("Запись «Неизвестных» не удалась", e))
        return
    send(chat, f"✅ В лист «Неизвестные устройства» дописано {n} строк "
               f"(бэкап {esc(bak.rsplit(chr(92), 1)[-1])})."
               f"\nПросмотр: /unknown · сверка: /reconcile unknown")


def write_unknowns():
    """Дозапись очереди в лист «Неизвестные устройства» с автобэкапом.
    -> (сколько строк, путь бэкапа). Существующие MAC листа не дублируются."""
    import openpyxl
    import bot_dq
    q = queue()
    have = _existing_unk_macs()
    rows = []
    now_s = time.strftime("%d.%m.%Y %H:%M")
    for k, e in sorted(q.items(), key=lambda kv: kv[1].get("ip") or ""):
        mac = e.get("mac") or ""
        if mac and inv.norm_mac(mac) in have:
            continue
        swp = (inv.switch_ports(mac) or [None])[0] if mac else None
        rows.append([mac or "?", net.vendor(mac) if mac else "?",
                     "не определено (бот)",
                     (swp or {}).get("host") or "", (swp or {}).get("sw_ip") or "",
                     (swp or {}).get("port") or "", (swp or {}).get("vlan") or "",
                     f"IP {e.get('ip')} · замечен ботом {now_s} "
                     f"({e.get('src') or '?'}, {e.get('seen', 1)}×)"])
    if not rows:
        store.jsave(_qpath(), {})
        return 0, "не потребовался"
    bak = bot_dq.backup_xlsx("unknown")
    path = inv.inv_path()
    wb = openpyxl.load_workbook(path)
    try:
        if inv.UNK_SHEET in wb.sheetnames:
            ws = wb[inv.UNK_SHEET]
            hdr = [str(c.value or "") for c in ws[1]]
        else:
            ws = wb.create_sheet(inv.UNK_SHEET)
            for i, h in enumerate(UNK_HDRS, 1):
                ws.cell(row=1, column=i, value=h)
            hdr = list(UNK_HDRS)
        idx = [hdr.index(h) if h in hdr else None for h in UNK_HDRS]
        rn = ws.max_row
        for row in rows:
            rn += 1
            for v, ci in zip(row, idx):
                if ci is not None and v not in (None, ""):
                    ws.cell(row=rn, column=ci + 1, value=v)
        wb.save(path)
    finally:
        wb.close()
    with inv._lock:
        inv._unk["mtime"] = None
    store.jsave(_qpath(), {})     # очередь записана — очищаем
    try:
        import bot_reconcile
        bot_reconcile.record_change("unknownq", inv.UNK_SHEET, f"{len(rows)} строк",
                                    "append", "", f"{len(rows)} хостов")
        bot_reconcile.after_xlsx_write(f"unknownq: {len(rows)} строк")
    except Exception:
        log_exc("unknownq: журнал изменений")
    try:
        import bot_obs
        bot_obs.audit("unknownq:write", f"{len(rows)} строк", "OK")
    except Exception:
        pass
    log(f"unknownq: записано {len(rows)} строк в «{inv.UNK_SHEET}», бэкап {bak}")
    return len(rows), bak


def _tick():
    """Раз в сутки в unknown_daily_hour — сводка владельцу с кнопкой."""
    if datetime.datetime.now().hour < int(st.cget("unknown_daily_hour")):
        return
    today = datetime.date.today().isoformat()
    s = store.jload(_qpath(), {})
    if (s.get("_meta") or {}).get("daily") == today:
        return
    store.jupdate(_qpath(), {}, lambda d: {**d, "_meta": {"daily": today}})
    q = queue()
    if not q:
        return
    owner = st.cget("owner_chat_id")
    if owner:
        _offer(owner, q)


HANDLERS = {"/unknown_queue": cmd_unknown_queue}
ALIASES = {"/uq": "/unknown_queue", "/очередь": "/unknown_queue"}
CALLBACKS = {"unkw": cb_unkw, "unkwy": cb_unkwy}

try:  # суточный тик в существующем health-цикле
    import bot_health as _bh
    _bh.MINUTE_TICKS.append(_tick)
except Exception:
    log_exc("unknownq: тик не зарегистрировался")
